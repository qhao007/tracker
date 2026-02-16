"""
导入导出功能 API 测试

测试导入导出 API 的功能：
- 模板下载 API
- 数据导入 API
- 数据导出 API

运行命令:
    cd dev && PYTHONPATH=. pytest tests/test_api/test_api_import_export.py -v
"""

import json
import pytest
import sys
import os
import time
import base64
import io
from openpyxl import Workbook

# 确保导入路径正确
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from app import create_app


@pytest.fixture
def client():
    """创建测试客户端"""
    app = create_app(testing=True)
    app.config['TESTING'] = True
    with app.test_client() as client:
        yield client


@pytest.fixture(scope='module')
def test_project():
    """创建测试项目用于测试"""
    app = create_app(testing=True)
    with app.test_client() as client:
        name = f"API_Test_{int(time.time())}"
        
        # 创建项目
        response = client.post('/api/projects',
                              data=json.dumps({'name': name}),
                              content_type='application/json')
        
        if response.status_code == 200:
            data = json.loads(response.data)
            project_id = data['project']['id']
            yield {'id': project_id, 'name': name}
            
            # 清理：删除测试项目
            client.delete(f"/api/projects/{project_id}")
        else:
            pytest.skip("无法创建测试项目")


class TestImportTemplateAPI:
    """测试导入模板下载 API"""
    
    def test_get_cp_template(self, client, test_project):
        """测试下载 CP 导入模板"""
        response = client.get('/api/import/template?type=cp')
        assert response.status_code == 200
        assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'CP_import_template.xlsx' in response.headers.get('Content-Disposition', '')
    
    def test_get_tc_template(self, client, test_project):
        """测试下载 TC 导入模板"""
        response = client.get('/api/import/template?type=tc')
        assert response.status_code == 200
        assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'TC_import_template.xlsx' in response.headers.get('Content-Disposition', '')
    
    def test_get_template_default_type(self, client, test_project):
        """测试默认模板类型（应为 CP）"""
        response = client.get('/api/import/template')
        assert response.status_code == 200
        assert 'CP_import_template.xlsx' in response.headers.get('Content-Disposition', '')
    
    def test_get_template_invalid_type(self, client, test_project):
        """测试无效的模板类型"""
        response = client.get('/api/import/template?type=invalid')
        assert response.status_code == 200  # 返回默认 CP 模板


class TestImportAPI:
    """测试数据导入 API"""
    
    def test_import_cp_success(self, client, test_project):
        """测试成功导入 CP"""
        # 创建 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.append(['Feature', 'Sub-Feature', 'Cover Point', 'Cover Point Details', 'Comments'])
        ws.append(['Feature1', 'SubFeature1', 'CP_Import_Test', 'Test CP', 'Test comment'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read()).decode('utf-8')
        
        response = client.post('/api/import', json={
            'project_id': test_project['id'],
            'type': 'cp',
            'file_data': file_content
        })
        
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['imported'] == 1
    
    def test_import_tc_success(self, client, test_project):
        """测试成功导入 TC"""
        # 创建 Excel 文件
        wb = Workbook()
        ws = wb.active
        ws.append(['TestBench', 'Category', 'Owner', 'Test Name', 'Scenario Details', 'Checker Details', 'Coverage Details', 'Comments'])
        ws.append(['TB1', 'Category1', 'Owner1', 'TC_Import_Test', 'Scenario', 'Checker', 'Coverage', 'Test comment'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read()).decode('utf-8')
        
        response = client.post('/api/import', json={
            'project_id': test_project['id'],
            'type': 'tc',
            'file_data': file_content
        })
        
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['imported'] == 1
    
    def test_import_cp_missing_required_field(self, client, test_project):
        """测试导入 CP 缺少必填字段"""
        # 缺少必填字段 Cover Point
        wb = Workbook()
        ws = wb.active
        ws.append(['Feature', 'Sub-Feature', 'Cover Point', 'Cover Point Details', 'Comments'])
        ws.append(['Feature1', 'SubFeature1', '', 'Test CP', 'Test comment'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read()).decode('utf-8')
        
        response = client.post('/api/import', json={
            'project_id': test_project['id'],
            'type': 'cp',
            'file_data': file_content
        })
        
        assert response.status_code == 400
        data = response.get_json()
        assert 'error' in data
    
    def test_import_tc_missing_required_field(self, client, test_project):
        """测试导入 TC 缺少必填字段"""
        # 缺少必填字段 Test Name
        wb = Workbook()
        ws = wb.active
        ws.append(['TestBench', 'Category', 'Owner', 'Test Name', 'Scenario Details'])
        ws.append(['TB1', 'Category1', 'Owner1', '', 'Scenario'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read()).decode('utf-8')
        
        response = client.post('/api/import', json={
            'project_id': test_project['id'],
            'type': 'tc',
            'file_data': file_content
        })
        
        assert response.status_code == 400
        data = response.get_json()
        assert 'error' in data
    
    def test_import_cp_duplicate(self, client, test_project):
        """测试导入重复的 CP"""
        # 先创建一个 CP
        client.post('/api/cp', json={
            'project_id': test_project['id'],
            'feature': 'Feature1',
            'cover_point': 'CP_Duplicate_Test',
            'priority': 'P0'
        })
        
        # 尝试导入相同的 CP
        wb = Workbook()
        ws = wb.active
        ws.append(['Feature', 'Sub-Feature', 'Cover Point', 'Cover Point Details', 'Comments'])
        ws.append(['Feature1', 'SubFeature1', 'CP_Duplicate_Test', 'Test CP', 'Test comment'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read()).decode('utf-8')
        
        response = client.post('/api/import', json={
            'project_id': test_project['id'],
            'type': 'cp',
            'file_data': file_content
        })
        
        assert response.status_code == 200
        data = response.get_json()
        assert data['imported'] == 0
        assert data['failed'] == 1
    
    def test_import_missing_params(self, client, test_project):
        """测试缺少必要参数"""
        response = client.post('/api/import', json={
            'project_id': test_project['id']
            # 缺少 type 和 file_data
        })
        
        assert response.status_code == 400
        data = response.get_json()
        assert 'error' in data
    
    def test_import_invalid_type(self, client, test_project):
        """测试无效的导入类型"""
        wb = Workbook()
        ws = wb.active
        ws.append(['Feature', 'Cover Point'])
        ws.append(['Feature1', 'CP_Test'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read()).decode('utf-8')
        
        response = client.post('/api/import', json={
            'project_id': test_project['id'],
            'type': 'invalid',
            'file_data': file_content
        })
        
        assert response.status_code == 400
        data = response.get_json()
        assert 'error' in data
    
    def test_import_invalid_project(self, client):
        """测试无效的项目 ID"""
        wb = Workbook()
        ws = wb.active
        ws.append(['Feature', 'Cover Point'])
        ws.append(['Feature1', 'CP_Test'])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = base64.b64encode(output.read()).decode('utf-8')
        
        response = client.post('/api/import', json={
            'project_id': 99999,
            'type': 'cp',
            'file_data': file_content
        })
        
        assert response.status_code == 404
        data = response.get_json()
        assert 'error' in data


class TestExportAPI:
    """测试数据导出 API"""
    
    def test_export_cp_xlsx(self, client, test_project):
        """测试导出 CP 为 Excel"""
        # 先创建一些 CP
        client.post('/api/cp', json={
            'project_id': test_project['id'],
            'feature': 'Feature1',
            'cover_point': 'CP_Export_1',
            'priority': 'P0'
        })
        
        response = client.get(f'/api/export?project_id={test_project["id"]}&type=cp&format=xlsx')
        
        assert response.status_code == 200
        assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'CP_' in response.headers.get('Content-Disposition', '')
        assert '.xlsx' in response.headers.get('Content-Disposition', '')
    
    def test_export_tc_xlsx(self, client, test_project):
        """测试导出 TC 为 Excel"""
        # 先创建一些 TC
        client.post('/api/tc', json={
            'project_id': test_project['id'],
            'testbench': 'TB1',
            'test_name': 'TC_Export_1',
            'dv_milestone': 'DV1.0'
        })
        
        response = client.get(f'/api/export?project_id={test_project["id"]}&type=tc&format=xlsx')
        
        assert response.status_code == 200
        assert response.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'TC_' in response.headers.get('Content-Disposition', '')
        assert '.xlsx' in response.headers.get('Content-Disposition', '')
    
    def test_export_cp_csv(self, client, test_project):
        """测试导出 CP 为 CSV"""
        response = client.get(f'/api/export?project_id={test_project["id"]}&type=cp&format=csv')
        
        assert response.status_code == 200
        assert 'text/csv' in response.content_type
        assert 'CP_' in response.headers.get('Content-Disposition', '')
        assert '.csv' in response.headers.get('Content-Disposition', '')
    
    def test_export_tc_csv(self, client, test_project):
        """测试导出 TC 为 CSV"""
        response = client.get(f'/api/export?project_id={test_project["id"]}&type=tc&format=csv')
        
        assert response.status_code == 200
        assert 'text/csv' in response.content_type
        assert 'TC_' in response.headers.get('Content-Disposition', '')
        assert '.csv' in response.headers.get('Content-Disposition', '')
    
    def test_export_default_format(self, client, test_project):
        """测试默认导出格式（应为 xlsx）"""
        response = client.get(f'/api/export?project_id={test_project["id"]}&type=cp')
        
        assert response.status_code == 200
        assert '.xlsx' in response.headers.get('Content-Disposition', '')
    
    def test_export_missing_params(self, client, test_project):
        """测试缺少必要参数"""
        response = client.get(f'/api/export?project_id={test_project["id"]}')
        assert response.status_code == 400
        
        response = client.get(f'/api/export?type=cp')
        assert response.status_code == 400
    
    def test_export_invalid_type(self, client, test_project):
        """测试无效的导出类型"""
        response = client.get(f'/api/export?project_id={test_project["id"]}&type=invalid')
        assert response.status_code == 400
    
    def test_export_invalid_project(self, client):
        """测试无效的项目 ID"""
        response = client.get('/api/export?project_id=99999&type=cp')
        assert response.status_code == 404
    
    def test_export_empty_project(self, client, test_project):
        """测试导出空项目"""
        response = client.get(f'/api/export?project_id={test_project["id"]}&type=cp')
        
        assert response.status_code == 200
        # 空项目应该也能导出，只是没有数据


class TestCSVImport:
    """测试 CSV 导入"""
    
    def test_import_cp_csv(self, client, test_project):
        """测试导入 CP CSV 文件"""
        csv_content = "Feature,Sub-Feature,Cover Point,Cover Point Details,Comments\nFeature1,SubFeature1,CP_CSV_Test,Test CP,Test comment"
        file_content = base64.b64encode(csv_content.encode('utf-8')).decode('utf-8')
        
        response = client.post('/api/import', json={
            'project_id': test_project['id'],
            'type': 'cp',
            'file_data': file_content
        })
        
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['imported'] == 1
    
    def test_import_tc_csv(self, client, test_project):
        """测试导入 TC CSV 文件"""
        csv_content = "TestBench,Category,Owner,Test Name,Scenario Details,Checker Details,Coverage Details,Comments\nTB1,Category1,Owner1,TC_CSV_Test,Scenario,Checker,Coverage,Test comment"
        file_content = base64.b64encode(csv_content.encode('utf-8')).decode('utf-8')
        
        response = client.post('/api/import', json={
            'project_id': test_project['id'],
            'type': 'tc',
            'file_data': file_content
        })
        
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['imported'] == 1
    
    def test_import_multiple_rows_csv(self, client, test_project):
        """测试导入多行 CSV"""
        csv_content = "Feature,Sub-Feature,Cover Point,Cover Point Details\nFeature1,SubFeature1,CP_1,Test 1\nFeature1,SubFeature1,CP_2,Test 2\nFeature1,SubFeature1,CP_3,Test 3"
        file_content = base64.b64encode(csv_content.encode('utf-8')).decode('utf-8')
        
        response = client.post('/api/import', json={
            'project_id': test_project['id'],
            'type': 'cp',
            'file_data': file_content
        })
        
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert data['imported'] == 3
