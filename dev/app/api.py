"""
Tracker API 路由 - v0.3 独立数据库版本
"""

from flask import Blueprint, request, jsonify, send_from_directory, send_file, current_app, g, session
from datetime import datetime, timedelta, date
from urllib.parse import quote
from functools import wraps
import json
import os
import sqlite3

api = Blueprint("api", __name__)
auth_api = Blueprint("auth_api", __name__)

# 导入认证模块
from . import auth

# 项目列表文件
PROJECTS_FILE = "data/projects.json"

# Session 配置
SESSION_USER_KEY = "user_id"
SESSION_USERNAME_KEY = "username"
SESSION_ROLE_KEY = "role"


# ============ 访问控制装饰器 ============

def login_required(f):
    """要求登录的装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if SESSION_USER_KEY not in session:
            return jsonify({"error": "Unauthorized", "message": "请先登录"}), 401
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """要求管理员权限的装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if SESSION_USER_KEY not in session:
            return jsonify({"error": "Unauthorized", "message": "请先登录"}), 401
        if session.get(SESSION_ROLE_KEY) != 'admin':
            return jsonify({"error": "Forbidden", "message": "需要管理员权限"}), 403
        return f(*args, **kwargs)
    return decorated_function


def guest_required(f):
    """禁止访客访问的装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if SESSION_USER_KEY not in session:
            return jsonify({"error": "Unauthorized", "message": "请先登录"}), 401
        if session.get(SESSION_ROLE_KEY) == 'guest':
            return jsonify({"error": "Forbidden", "message": "访客无权限执行此操作"}), 403
        return f(*args, **kwargs)
    return decorated_function


def get_projects_file():
    """获取项目列表文件路径"""
    return os.path.join(current_app.config["DATA_DIR"], "projects.json")


def load_projects():
    """加载项目列表"""
    filepath = get_projects_file()
    if os.path.exists(filepath):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, FileNotFoundError, IOError):
            pass
    return []


def save_projects(projects):
    """保存项目列表"""
    filepath = get_projects_file()
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(projects, f, ensure_ascii=False, indent=2)


def get_db_path(project_name):
    """获取项目数据库文件路径"""
    safe_name = project_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
    return os.path.join(current_app.config["DATA_DIR"], f"{safe_name}.db")


def get_db(project_name):
    """获取项目数据库连接"""
    if "db_connections" not in g:
        g.db_connections = {}

    if project_name not in g.db_connections:
        db_path = get_db_path(project_name)
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        g.db_connections[project_name] = conn

    return g.db_connections[project_name]


def init_project_db(project_name):
    """初始化项目数据库"""
    db_path = get_db_path(project_name)
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # 创建 Cover Points 表
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS cover_point (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            feature TEXT,
            sub_feature TEXT,
            cover_point TEXT,
            cover_point_details TEXT,
            comments TEXT,
            priority TEXT DEFAULT 'P0',
            created_at TEXT
        )
    """)

    # 创建 Test Cases 表
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS test_case (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            dv_milestone TEXT,
            testbench TEXT,
            category TEXT,
            owner TEXT,
            test_name TEXT,
            scenario_details TEXT,
            checker_details TEXT,
            coverage_details TEXT,
            comments TEXT,
            priority TEXT DEFAULT 'P0',
            status TEXT DEFAULT 'OPEN',
            created_at TEXT,
            coded_date TEXT,
            fail_date TEXT,
            pass_date TEXT,
            removed_date TEXT,
            target_date TEXT
        )
    """)

    # 创建关联表 (兼容两种结构: 旧版自增ID / 新版复合主键)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tc_cp_connections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tc_id INTEGER,
            cp_id INTEGER,
            UNIQUE(tc_id, cp_id)
        )
    """)

    # 创建 Functional Coverage 表 (v0.11.0)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS functional_coverage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            covergroup TEXT NOT NULL,
            coverpoint TEXT NOT NULL,
            coverage_type TEXT NOT NULL,
            bin_name TEXT NOT NULL,
            bin_val TEXT,
            comments TEXT,
            coverage_pct REAL DEFAULT 0.0,
            status TEXT DEFAULT 'missing' CHECK (status IN ('missing', 'ready')),
            owner TEXT,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now')),
            UNIQUE (project_id, covergroup, coverpoint, bin_name)
        )
    """)

    # 创建 FC-CP 关联表 (v0.11.0)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS fc_cp_association (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            cp_id INTEGER,
            fc_id INTEGER,
            created_by TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            UNIQUE (cp_id, fc_id),
            FOREIGN KEY (cp_id) REFERENCES cover_point(id),
            FOREIGN KEY (fc_id) REFERENCES functional_coverage(id)
        )
    """)

    # 创建索引 (v0.11.0)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_fc_covergroup ON functional_coverage(covergroup)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_fc_coverpoint ON functional_coverage(coverpoint)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_fc_coverage_type ON functional_coverage(coverage_type)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_fc_cp_assoc_cp ON fc_cp_association(cp_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_fc_cp_assoc_fc ON fc_cp_association(fc_id)")

    conn.commit()
    conn.close()


def delete_project_db(project_name):
    """删除项目数据库文件"""
    db_path = get_db_path(project_name)
    if os.path.exists(db_path):
        os.remove(db_path)


def jsonify_project(project_name, data):
    """将项目数据转为列表格式"""
    return jsonify(data)


# ============ 版本信息 ============


@api.route("/api/version", methods=["GET"])
def get_version():
    """获取版本信息"""
    import os

    version_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "VERSION")
    version = "1.0.0"
    release_date = "2026-02-04"
    if os.path.exists(version_file):
        try:
            with open(version_file, "r") as f:
                lines = f.readlines()
                for line in lines:
                    line = line.strip()
                    if "=" in line:
                        key, value = line.split("=", 1)
                        if key == "VERSION":
                            version = value
                        elif key == "RELEASE_DATE":
                            release_date = value
                    elif line:
                        # 兼容只有版本号的格式，如 "v0.6.1"
                        version = line
        except (json.JSONDecodeError, IndexError, KeyError):
            pass
    return jsonify({"version": version, "version_type": "正式版", "release_date": release_date})


# ============ 项目管理 ============

@api.route("/api/projects", methods=["GET"])
@login_required
def get_projects():
    """获取项目列表"""
    projects = load_projects()
    result = []
    for p in projects:
        if p.get("is_archived", False):
            continue

        # 统计项目数据
        try:
            db = get_db(p["name"])
            cursor = db.cursor()
            cursor.execute("SELECT COUNT(*) FROM cover_point")
            cp_count = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM test_case")
            tc_count = cursor.fetchone()[0]
        except (sqlite3.Error, KeyError):
            cp_count = 0
            tc_count = 0

        result.append(
            {
                "id": p["id"],
                "name": p["name"],
                "created_at": p.get("created_at", ""),
                "is_archived": False,
                "version": p.get("version", "stable"),
                "start_date": p.get("start_date", ""),
                "end_date": p.get("end_date", ""),
                "coverage_mode": p.get("coverage_mode", "tc_cp"),  # v0.11.0
                "cp_count": cp_count,
                "tc_count": tc_count,
            }
        )

    return jsonify(result)


@api.route("/api/projects/<int:project_id>", methods=["GET"])
def get_project(project_id):
    """获取项目详情"""
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    # 如果项目已归档，返回错误
    if project.get("is_archived", False):
        return jsonify({"error": "项目已归档"}), 404

    # 统计项目数据
    try:
        db = get_db(project["name"])
        cursor = db.cursor()
        cursor.execute("SELECT COUNT(*) FROM cover_point")
        cp_count = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM test_case")
        tc_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM test_case WHERE status = "PASS"')
        pass_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM test_case WHERE status = "REMOVED"')
        removed_count = cursor.fetchone()[0]
    except sqlite3.Error:
        cp_count = 0
        tc_count = 0
        pass_count = 0
        removed_count = 0

    return jsonify(
        {
            "id": project["id"],
            "name": project["name"],
            "created_at": project.get("created_at", ""),
            "is_archived": project.get("is_archived", False),
            "version": project.get("version", "stable"),
            "start_date": project.get("start_date", ""),
            "end_date": project.get("end_date", ""),
            "coverage_mode": project.get("coverage_mode", "tc_cp"),  # v0.11.0
            "cp_count": cp_count,
            "tc_count": tc_count,
            "pass_count": pass_count,
            "removed_count": removed_count,
        }
    )


@api.route("/api/projects", methods=["POST"])
@admin_required
def create_project():
    """创建新项目"""
    data = request.json
    name = data.get("name", "").strip()
    start_date = data.get("start_date", "").strip()
    end_date = data.get("end_date", "").strip()
    create_test_user = data.get("create_test_user", False)  # v0.8.3: 自动创建测试用户
    # v0.11.0: coverage_mode - 'tc_cp' 或 'fc_cp'
    coverage_mode = data.get("coverage_mode", "tc_cp").strip()

    if not name:
        return jsonify({"error": "项目名称不能为空"}), 400

    # v0.8.3: 日期必填验证
    if not start_date:
        return jsonify({"error": "开始日期不能为空"}), 400
    if not end_date:
        return jsonify({"error": "结束日期不能为空"}), 400

    # 校验日期格式
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
            if start > end:
                return jsonify({"error": "开始日期不能晚于结束日期"}), 400
        except ValueError:
            return jsonify({"error": "日期格式应为 YYYY-MM-DD"}), 400

    projects = load_projects()
    if any(p["name"] == name and not p.get("is_archived", False) for p in projects):
        return jsonify({"error": f'项目 "{name}" 已存在'}), 400

    # 创建项目数据库
    init_project_db(name)

    # 添加到项目列表
    project = {
        "id": max([p["id"] for p in projects], default=0) + 1,
        "name": name,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "is_archived": False,
        "version": "stable",
        "start_date": start_date,
        "end_date": end_date,
        "coverage_mode": coverage_mode,  # v0.11.0
    }
    projects.append(project)
    save_projects(projects)

    # v0.8.3: 如果需要，自动创建测试用户
    test_user_info = None
    if create_test_user:
        # 生成测试用户名
        test_username = f"test_user_{name.replace(' ', '_').lower()}"
        test_password = "test_user123"
        
        # 检查用户是否已存在
        existing = auth.get_user_by_username(test_username)
        if not existing:
            try:
                user_id = auth.create_user(test_username, test_password, "user")
                test_user_info = {
                    "username": test_username,
                    "password": test_password,
                    "role": "user"
                }
            except Exception as e:
                # 用户创建失败不影响项目创建
                pass
        else:
            # 用户已存在
            test_user_info = {
                "username": test_username,
                "password": test_password,
                "role": "user",
                "note": "用户已存在"
            }

    response = {"success": True, "project": project}
    if test_user_info:
        response["test_user"] = test_user_info
    
    return jsonify(response)


@api.route("/api/projects/<int:project_id>", methods=["PUT"])
@admin_required
def update_project(project_id):
    """更新项目信息"""
    data = request.json
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    # 校验并更新字段
    if "name" in data:
        new_name = data["name"].strip()
        if new_name and new_name != project["name"]:
            # 检查名称是否已存在
            if any(p["name"] == new_name and p["id"] != project_id for p in projects):
                return jsonify({"error": f'项目名称 "{new_name}" 已存在'}), 400
            project["name"] = new_name

    # 更新日期字段（v0.8.0）
    if "start_date" in data:
        project["start_date"] = data["start_date"].strip() if data["start_date"] else ""
    if "end_date" in data:
        project["end_date"] = data["end_date"].strip() if data["end_date"] else ""

    # 校验日期（如果提供了日期）
    start_date = project.get("start_date", "")
    end_date = project.get("end_date", "")
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            end = datetime.strptime(end_date, "%Y-%m-%d")
            if start > end:
                return jsonify({"error": "开始日期不能晚于结束日期"}), 400
        except ValueError:
            return jsonify({"error": "日期格式应为 YYYY-MM-DD"}), 400

    # 保存更新
    save_projects(projects)

    return jsonify({"success": True, "project": project})


@api.route("/api/projects/<int:project_id>/archive", methods=["POST"])
@admin_required
def archive_project(project_id):
    """备份项目"""
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    # 收集项目数据
    conn = get_db(project["name"])
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM cover_point")
    cps = [dict(row) for row in cursor.fetchall()]

    cursor.execute("SELECT * FROM test_case")
    tcs = [dict(row) for row in cursor.fetchall()]

    project_data = {
        "id": project["id"],
        "name": project["name"],
        "created_at": project.get("created_at", ""),
        "version": project.get("version", "stable"),
        "cover_points": cps,
        "test_cases": tcs,
        "backup_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    # 生成备份文件
    filename = f"{project['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    archives_dir = "archives"
    os.makedirs(archives_dir, exist_ok=True)  # 确保 archives 目录存在
    filepath = os.path.join(archives_dir, filename)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(project_data, f, ensure_ascii=False, indent=2)

    return jsonify({"success": True, "filename": filename})


@api.route("/api/projects/archive/list", methods=["GET"])
@login_required
def list_archives():
    """获取归档列表"""
    archives_dir = "archives"
    if not os.path.exists(archives_dir):
        return jsonify([])

    archives = []
    for f in os.listdir(archives_dir):
        if f.endswith(".json") and not f.startswith("ArchiveTest"):
            filepath = os.path.join(archives_dir, f)
            try:
                with open(filepath, "r", encoding="utf-8") as file:
                    data = json.load(file)
                    archives.append(
                        {
                            "id": len(archives) + 1,
                            "filename": f,
                            "project_name": data.get("name", f),
                            "backup_date": data.get("backup_date", ""),
                        }
                    )
            except (json.JSONDecodeError, KeyError, TypeError):
                pass

    return jsonify(archives)


@api.route("/api/projects/restore", methods=["POST"])
def restore_project():
    """从归档恢复项目"""
    data = request.json
    filename = data.get("filename")

    if not filename:
        return jsonify({"error": "需要指定备份文件名"}), 400

    filepath = f"archives/{filename}"
    if not os.path.exists(filepath):
        return jsonify({"error": "备份文件不存在"}), 404

    with open(filepath, "r", encoding="utf-8") as f:
        project_data = json.load(f)

    project_name = project_data.get("name")

    # 检查项目是否已存在
    projects = load_projects()
    if any(p["name"] == project_name and not p.get("is_archived", False) for p in projects):
        return jsonify({"error": f'项目 "{project_name}" 已存在，无法恢复'}), 400

    # 初始化新项目数据库
    init_project_db(project_name)
    conn = get_db(project_name)
    cursor = conn.cursor()

    # 恢复 Cover Points
    for cp_data in project_data.get("cover_points", []):
        cursor.execute(
            """
            INSERT INTO cover_point (project_id, feature, sub_feature, cover_point, cover_point_details, comments, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
            (
                cp_data.get("project_id", project_data["id"]),
                cp_data.get("feature", ""),
                cp_data.get("sub_feature", ""),
                cp_data.get("cover_point", ""),
                cp_data.get("cover_point_details", ""),
                cp_data.get("comments", ""),
                cp_data.get("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ),
        )

    # 恢复 Test Cases
    for tc_data in project_data.get("test_cases", []):
        cursor.execute(
            """
            INSERT INTO test_case (project_id, dv_milestone, priority, testbench, category, owner, test_name, scenario_details, checker_details, coverage_details, comments, status, completed_date, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                tc_data.get("project_id", project_data["id"]),
                tc_data.get("dv_milestone", ""),
                tc_data.get("priority", ""),
                tc_data.get("testbench", ""),
                tc_data.get("category", ""),
                tc_data.get("owner", ""),
                tc_data.get("test_name", ""),
                tc_data.get("scenario_details", ""),
                tc_data.get("checker_details", ""),
                tc_data.get("coverage_details", ""),
                tc_data.get("comments", ""),
                tc_data.get("status", "OPEN"),
                tc_data.get("completed_date", ""),
                tc_data.get("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ),
        )

    conn.commit()

    # 添加到项目列表
    project = {
        "id": max([p["id"] for p in projects], default=0) + 1,
        "name": project_name,
        "created_at": project_data.get("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        "is_archived": False,
        "version": project_data.get("version", "stable"),
    }
    projects.append(project)
    save_projects(projects)

    return jsonify({"success": True, "project": project})


@api.route("/api/projects/restore/upload", methods=["POST"])
def restore_project_upload():
    """从上传文件恢复项目"""
    # 检查是否有文件
    if "file" not in request.files:
        return jsonify({"error": "没有文件"}), 400

    file = request.files["file"]

    # 检查文件名
    if file.filename == "":
        return jsonify({"error": "没有选择文件"}), 400

    # 验证是 JSON 文件
    if not file.filename.endswith(".json"):
        return jsonify({"error": "只支持 JSON 格式的备份文件"}), 400

    try:
        # 读取并解析 JSON
        project_data = json.load(file)
    except Exception as e:
        return jsonify({"error": f"解析 JSON 文件失败: {str(e)}"}), 400

    project_name = project_data.get("name")

    if not project_name:
        return jsonify({"error": "备份文件缺少项目名称"}), 400

    # 检查项目是否已存在
    projects = load_projects()
    if any(p["name"] == project_name and not p.get("is_archived", False) for p in projects):
        return jsonify({"error": f'项目 "{project_name}" 已存在，无法恢复'}), 400

    # 初始化新项目数据库
    init_project_db(project_name)
    conn = get_db(project_name)
    cursor = conn.cursor()

    # 恢复 Cover Points
    for cp_data in project_data.get("cover_points", []):
        cursor.execute(
            """
            INSERT INTO cover_point (project_id, feature, sub_feature, cover_point, cover_point_details, comments, priority, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                cp_data.get("project_id", project_data.get("id", 1)),
                cp_data.get("feature", ""),
                cp_data.get("sub_feature", ""),
                cp_data.get("cover_point", ""),
                cp_data.get("cover_point_details", ""),
                cp_data.get("comments", ""),
                cp_data.get("priority", "P0"),
                cp_data.get("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ),
        )

    # 恢复 Test Cases
    for tc_data in project_data.get("test_cases", []):
        cursor.execute(
            """
            INSERT INTO test_case (project_id, dv_milestone, testbench, category, owner, test_name, 
            scenario_details, checker_details, coverage_details, comments, status, created_at,
            coded_date, fail_date, pass_date, removed_date, target_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                tc_data.get("project_id", project_data.get("id", 1)),
                tc_data.get("dv_milestone", ""),
                tc_data.get("testbench", ""),
                tc_data.get("category", ""),
                tc_data.get("owner", ""),
                tc_data.get("test_name", ""),
                tc_data.get("scenario_details", ""),
                tc_data.get("checker_details", ""),
                tc_data.get("coverage_details", ""),
                tc_data.get("comments", ""),
                tc_data.get("status", "OPEN"),
                tc_data.get("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                tc_data.get("coded_date"),
                tc_data.get("fail_date"),
                tc_data.get("pass_date"),
                tc_data.get("removed_date"),
                tc_data.get("target_date"),
            ),
        )

    conn.commit()

    # 添加到项目列表
    project = {
        "id": max([p["id"] for p in projects], default=0) + 1,
        "name": project_name,
        "created_at": project_data.get("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        "is_archived": False,
        "version": project_data.get("version", "stable"),
    }
    projects.append(project)
    save_projects(projects)

    return jsonify({"success": True, "project": project})


@api.route("/api/projects/<int:project_id>", methods=["DELETE"])
@admin_required
def delete_project(project_id):
    """删除项目"""
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    # 删除前自动创建归档备份
    try:
        # 收集项目数据
        conn = get_db(project["name"])
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM cover_point")
        cps = [dict(row) for row in cursor.fetchall()]

        cursor.execute("SELECT * FROM test_case")
        tcs = [dict(row) for row in cursor.fetchall()]

        project_data = {
            "id": project["id"],
            "name": project["name"],
            "created_at": project.get("created_at", ""),
            "version": project.get("version", "stable"),
            "cover_points": cps,
            "test_cases": tcs,
            "backup_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        # 生成备份文件（保存在 archives 目录）
        filename = f"{project['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_deleted.json"
        archives_dir = "archives"
        os.makedirs(archives_dir, exist_ok=True)
        filepath = os.path.join(archives_dir, filename)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(project_data, f, ensure_ascii=False, indent=2)

    except Exception as e:
        # 备份失败不影响删除流程，但记录错误
        print(f"项目备份失败: {e}")

    # 标记为归档
    project["is_archived"] = True
    save_projects(projects)

    # 删除数据库文件
    delete_project_db(project["name"])

    return jsonify({"success": True})


# ============ Progress Charts (v0.8.0) ============


def calculate_planned_coverage(project_name, start_date, end_date, priority_filter=""):
    """
    计算计划覆盖率曲线

    算法：对于项目周期内的每一周，计算该周之前已target且状态为Pass的TC
    所关联的CP覆盖率

    Args:
        project_name: 项目名称
        start_date: 项目开始日期
        end_date: 项目结束日期
        priority_filter: CP priority 过滤参数，多值用逗号分隔，如 "P0,P1"

    Returns:
        list: 每周数据点列表 [{week: 'YYYY-MM-DD', coverage: float}, ...]
    """
    if not start_date or not end_date:
        return []

    try:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return []

    # 获取项目 ID
    projects = load_projects()
    project = next((p for p in projects if p["name"] == project_name), None)
    if not project:
        return []
    project_id = project["id"]

    conn = get_db(project_name)
    cursor = conn.cursor()

    # 构建 priority 过滤条件 (v0.10.0)
    priority_condition = ""
    priority_params = []
    if priority_filter:
        priority_list = [p.strip() for p in priority_filter.split(',') if p.strip()]
        if priority_list:
            placeholders = ','.join(['?' for _ in priority_list])
            priority_condition = f"AND cp.priority IN ({placeholders})"
            priority_params = priority_list

    # 获取项目总 CP 数（根据过滤条件）
    if priority_condition:
        cursor.execute(f"SELECT COUNT(*) FROM cover_point cp WHERE cp.project_id = ? {priority_condition}", [project_id] + priority_params)
    else:
        cursor.execute("SELECT COUNT(*) FROM cover_point WHERE project_id = ?", (project_id,))
    total_cp = cursor.fetchone()[0]

    if total_cp == 0:
        return []

    planned = []

    # 从项目开始日期到结束日期，每周计算一次
    current = start
    while current <= end:
        # 获取本周的周日（week end）
        # 周一为一周开始
        days_to_sunday = (current.weekday() - 6) % 7  # 0-6, Monday=0, Sunday=6
        if days_to_sunday == 0 and current.weekday() == 6:
            # 如果当前是周日，直接用
            week_end = current
        else:
            # 计算本周日的日期
            days_until_sunday = (6 - current.weekday()) % 7
            if days_until_sunday == 0:
                days_until_sunday = 7
            week_end = current + timedelta(days=days_until_sunday)

        # 只处理不超过结束日期的周
        if week_end > end:
            week_end = end

        # 计算覆盖率：target_date <= week_end 且 status = 'Pass' 的 TC 关联的 CP 覆盖率
        # 支持 priority 过滤 (v0.10.0)
        if priority_condition:
            cursor.execute(f"""
                WITH pass_tcs AS (
                    SELECT DISTINCT tc.id
                    FROM test_case tc
                    WHERE tc.project_id = ?
                    AND tc.target_date IS NOT NULL
                    AND tc.target_date <= ?
                    AND tc.status != 'REMOVED'
                ),
                covered_cps AS (
                    SELECT DISTINCT cp.id
                    FROM tc_cp_connections tcc
                    INNER JOIN pass_tcs pt ON tcc.tc_id = pt.id
                    INNER JOIN cover_point cp ON tcc.cp_id = cp.id
                    WHERE 1=1 {priority_condition}
                )
                SELECT COUNT(*) FROM covered_cps
            """, [project_id, week_end.isoformat()] + priority_params)
        else:
            cursor.execute("""
                WITH pass_tcs AS (
                    SELECT DISTINCT tc.id
                    FROM test_case tc
                    WHERE tc.project_id = ?
                    AND tc.target_date IS NOT NULL
                    AND tc.target_date <= ?
                    AND tc.status != 'REMOVED'
                ),
                covered_cps AS (
                    SELECT DISTINCT cp.id
                    FROM tc_cp_connections tcc
                    INNER JOIN pass_tcs pt ON tcc.tc_id = pt.id
                    INNER JOIN cover_point cp ON tcc.cp_id = cp.id
                )
                SELECT COUNT(*) FROM covered_cps
            """, (project_id, week_end.isoformat()))

        covered_cp = cursor.fetchone()[0]
        coverage = round((covered_cp / total_cp) * 100, 1)

        planned.append({
            'week': current.isoformat(),  # 使用周一的日期作为标识
            'coverage': coverage
        })

        # 移到下一周
        current = week_end + timedelta(days=1)

    return planned


@api.route("/api/progress/<int:project_id>", methods=["GET"])
@login_required
def get_progress(project_id):
    """获取项目进度数据"""
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    # 获取项目日期
    start_date = project.get("start_date", "")
    end_date = project.get("end_date", "")

    # 获取日期过滤参数
    filter_start = request.args.get("start_date", "")
    filter_end = request.args.get("end_date", "")

    # 如果有过滤参数，使用过滤参数
    if filter_start:
        start_date = filter_start
    if filter_end:
        end_date = filter_end

    # 获取 priority 过滤参数 (v0.10.0)
    priority_filter = request.args.get("priority", "")

    # 计算计划曲线
    planned = calculate_planned_coverage(project["name"], start_date, end_date, priority_filter)

    # v0.8.1: 返回计划曲线
    # v0.8.2: 添加实际曲线
    # v0.10.0: 支持 Priority 过滤
    actual = []
    # 确保表存在并获取快照
    try:
        ensure_progress_table_exists(project["name"])
        conn = get_db(project["name"])
        cursor = conn.cursor()

        # v0.10.0: 根据 priority 参数选择覆盖率字段
        # v0.10.0: 修复多值 Priority 处理 - 计算加权平均覆盖率
        ALLOWED_COVERAGE_COLUMNS = {'actual_coverage', 'p0_coverage', 'p1_coverage', 'p2_coverage', 'p3_coverage'}
        coverage_column = 'actual_coverage'
        use_weighted_average = False

        if priority_filter:
            # 解析 priority 参数
            priority_list = [p.strip().upper() for p in priority_filter.split(',') if p.strip()]
            # 过滤掉无效值
            priority_list = [p for p in priority_list if p in ('P0', 'P1', 'P2', 'P3')]

            if len(priority_list) == 1:
                # 单个 Priority：直接使用该 Priority 的覆盖率
                first_priority = priority_list[0].lower()
                candidate_column = f'{first_priority}_coverage'
                if candidate_column in ALLOWED_COVERAGE_COLUMNS:
                    coverage_column = candidate_column
            elif len(priority_list) > 1:
                # 多个 Priority：检查是否覆盖了全部 CP
                total_project_cp = cursor.execute(
                    "SELECT COUNT(*) FROM cover_point WHERE project_id = ?",
                    (project_id,)
                ).fetchone()[0]

                priority_cp_counts = {}
                total_filtered_cp = 0
                for p in ('P0', 'P1', 'P2', 'P3'):
                    cursor.execute(
                        "SELECT COUNT(*) FROM cover_point WHERE project_id = ? AND priority = ?",
                        (project_id, p)
                    )
                    count = cursor.fetchone()[0]
                    priority_cp_counts[p] = count
                    if p in priority_list:
                        total_filtered_cp += count

                # v0.10.0: 如果覆盖了全部 CP，使用 actual_coverage
                if total_filtered_cp >= total_project_cp:
                    coverage_column = 'actual_coverage'
                    use_weighted_average = False
                else:
                    # 部分覆盖，使用加权平均
                    use_weighted_average = True

        if use_weighted_average:
            # 从数据库获取每个 Priority 的 CP 数量
            priority_cp_counts = {}
            total_filtered_cp = 0
            for p in ('P0', 'P1', 'P2', 'P3'):
                if f'{p.lower()}_coverage' in ALLOWED_COVERAGE_COLUMNS:
                    cursor.execute(
                        "SELECT COUNT(*) FROM cover_point WHERE project_id = ? AND priority = ?",
                        (project_id, p)
                    )
                    count = cursor.fetchone()[0]
                    priority_cp_counts[p] = count
                    total_filtered_cp += count

            # 获取快照数据并计算加权平均
            cursor.execute("""
                SELECT snapshot_date, actual_coverage, p0_coverage, p1_coverage, p2_coverage, p3_coverage
                FROM project_progress
                WHERE project_id = ?
                ORDER BY snapshot_date
            """, (project_id,))
            for row in cursor.fetchall():
                week = row[0]
                # 计算加权平均
                weighted_sum = 0
                if total_filtered_cp > 0:
                    coverage_cols = [row[2], row[3], row[4], row[5]]  # p0, p1, p2, p3
                    priorities = ['P0', 'P1', 'P2', 'P3']
                    for i, p in enumerate(priorities):
                        if priority_cp_counts[p] > 0 and coverage_cols[i] is not None:
                            weighted_sum += coverage_cols[i] * priority_cp_counts[p]
                    coverage = round(weighted_sum / total_filtered_cp, 1)
                else:
                    coverage = row[1] if row[1] is not None else 0
                actual.append({
                    'week': week,
                    'coverage': coverage
                })
        else:
            cursor.execute(f"""
                SELECT snapshot_date, {coverage_column}
                FROM project_progress
                WHERE project_id = ?
                ORDER BY snapshot_date
            """, (project_id,))
            for row in cursor.fetchall():
                actual.append({
                    'week': row[0],
                    'coverage': row[1] if row[1] is not None else 0
                })
    except Exception as e:
        print(f"Warning: Could not load actual curve: {e}")

    return jsonify({
        "project_id": project_id,
        "project_name": project["name"],
        "start_date": project.get("start_date", ""),
        "end_date": project.get("end_date", ""),
        "planned": planned,
        "actual": actual,
        "priority_filter": priority_filter
    })


# ============ Progress Snapshots (v0.8.2) ============


def ensure_progress_table_exists(project_name):
    """
    确保 project_progress 表存在，如不存在则创建
    v0.10.0: 扩展支持 p0_coverage~p3_coverage 字段

    Returns:
        bool: 表是否存在或创建成功
    """
    conn = get_db(project_name)
    cursor = conn.cursor()

    # 检查表是否存在
    cursor.execute("""
        SELECT name FROM sqlite_master
        WHERE type='table' AND name='project_progress'
    """)

    if cursor.fetchone():
        # 表已存在，检查是否需要添加 Priority 覆盖率字段 (v0.10.0)
        cursor.execute("PRAGMA table_info(project_progress)")
        columns = [row[1] for row in cursor.fetchall()]

        # 如果缺少 Priority 覆盖率字段，则添加
        for priority in ['p0', 'p1', 'p2', 'p3']:
            col_name = f'{priority}_coverage'
            if col_name not in columns:
                try:
                    cursor.execute(f"ALTER TABLE project_progress ADD COLUMN {col_name} REAL")
                    conn.commit()
                except Exception as e:
                    print(f"Warning: Could not add column {col_name}: {e}")
        return True

    # 创建表 (包含 Priority 覆盖率字段)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS project_progress (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            snapshot_date TEXT NOT NULL,
            actual_coverage REAL,
            p0_coverage REAL,
            p1_coverage REAL,
            p2_coverage REAL,
            p3_coverage REAL,
            tc_pass_count INTEGER,
            tc_total INTEGER,
            cp_covered INTEGER,
            cp_total INTEGER,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT,
            updated_by TEXT,
            UNIQUE(project_id, snapshot_date)
        )
    """)
    conn.commit()
    return True


def calculate_current_coverage(project_name):
    """
    计算当前覆盖率（用于快照）
    v0.10.0: 扩展支持各 Priority 覆盖率计算

    Returns:
        dict: {actual_coverage, p0_coverage, p1_coverage, p2_coverage, p3_coverage,
               tc_pass_count, tc_total, cp_covered, cp_total}
    """
    # 确保 project_progress 表存在
    try:
        ensure_progress_table_exists(project_name)
    except Exception:
        return None

    # 获取项目 ID
    projects = load_projects()
    project = next((p for p in projects if p["name"] == project_name), None)
    if not project:
        return None

    conn = get_db(project_name)
    cursor = conn.cursor()

    # 获取总 CP 数
    cursor.execute("SELECT COUNT(*) FROM cover_point")
    total_cp = cursor.fetchone()[0]

    if total_cp == 0:
        return {
            'actual_coverage': 0,
            'p0_coverage': 0,
            'p1_coverage': 0,
            'p2_coverage': 0,
            'p3_coverage': 0,
            'tc_pass_count': 0,
            'tc_total': 0,
            'cp_covered': 0,
            'cp_total': total_cp
        }

    # 获取 Pass 状态的 TC
    cursor.execute("SELECT COUNT(*) FROM test_case WHERE status = 'PASS'")
    tc_pass = cursor.fetchone()[0]

    # 获取总 TC 数
    cursor.execute("SELECT COUNT(*) FROM test_case")
    tc_total = cursor.fetchone()[0]

    # 获取 Pass TC 关联的 CP（去重）
    cursor.execute("""
        SELECT DISTINCT cp.id
        FROM test_case tc
        INNER JOIN tc_cp_connections tcc ON tc.id = tcc.tc_id
        INNER JOIN cover_point cp ON tcc.cp_id = cp.id
        WHERE tc.status != 'REMOVED'
    """)
    results = cursor.fetchall()
    covered_cps = len(results) if results else 0

    coverage = round((covered_cps / total_cp) * 100, 1) if total_cp > 0 else 0

    # v0.10.0: 计算各 Priority 的覆盖率
    priority_coverages = {}
    for priority in ['P0', 'P1', 'P2', 'P3']:
        # 获取该 Priority 的总 CP 数
        cursor.execute(
            "SELECT COUNT(*) FROM cover_point WHERE priority = ?",
            (priority,)
        )
        priority_total = cursor.fetchone()[0]

        if priority_total == 0:
            priority_coverages[f'{priority.lower()}_coverage'] = 0
            continue

        # 获取该 Priority 下被 Pass TC 覆盖的 CP 数
        cursor.execute("""
            SELECT DISTINCT cp.id
            FROM test_case tc
            INNER JOIN tc_cp_connections tcc ON tc.id = tcc.tc_id
            INNER JOIN cover_point cp ON tcc.cp_id = cp.id
            WHERE tc.status != 'REMOVED' AND cp.priority = ?
        """, (priority,))
        results = cursor.fetchall()
        priority_covered = len(results) if results else 0

        priority_coverage = round((priority_covered / priority_total) * 100, 1) if priority_total > 0 else 0
        priority_coverages[f'{priority.lower()}_coverage'] = priority_coverage

    return {
        'actual_coverage': coverage,
        'tc_pass_count': tc_pass,
        'tc_total': tc_total,
        'cp_covered': covered_cps,
        'cp_total': total_cp,
        **priority_coverages
    }


@api.route("/api/progress/<int:project_id>/snapshot", methods=["POST"])
@admin_required
def create_snapshot(project_id):
    """手动创建进度快照"""
    from datetime import datetime
    
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    
    if not project:
        return jsonify({"error": "项目不存在"}), 404
    
    # 确保表存在
    ensure_progress_table_exists(project["name"])
    
    # 计算当前覆盖率
    coverage_data = calculate_current_coverage(project["name"])
    if coverage_data is None:
        return jsonify({"error": "数据库未初始化"}), 500
    
    today = datetime.now().strftime('%Y-%m-%d')
    conn = get_db(project["name"])
    cursor = conn.cursor()
    
    # 检查是否已存在当周快照
    cursor.execute("""
        SELECT id FROM project_progress 
        WHERE project_id = ? AND snapshot_date = ?
    """, (project_id, today))
    existing = cursor.fetchone()
    
    if existing:
        # 更新现有快照 (v0.10.0: 包含 Priority 覆盖率)
        cursor.execute("""
            UPDATE project_progress SET
                actual_coverage = ?,
                p0_coverage = ?,
                p1_coverage = ?,
                p2_coverage = ?,
                p3_coverage = ?,
                tc_pass_count = ?,
                tc_total = ?,
                cp_covered = ?,
                cp_total = ?,
                updated_at = ?,
                updated_by = ?
            WHERE project_id = ? AND snapshot_date = ?
        """, (
            coverage_data['actual_coverage'],
            coverage_data.get('p0_coverage', 0),
            coverage_data.get('p1_coverage', 0),
            coverage_data.get('p2_coverage', 0),
            coverage_data.get('p3_coverage', 0),
            coverage_data['tc_pass_count'],
            coverage_data['tc_total'],
            coverage_data['cp_covered'],
            coverage_data['cp_total'],
            today,
            session.get('username', 'admin'),
            project_id,
            today
        ))
        conn.commit()
        snapshot_id = existing[0]
    else:
        # 创建新快照 (v0.10.0: 包含 Priority 覆盖率)
        cursor.execute("""
            INSERT INTO project_progress (
                project_id, snapshot_date, actual_coverage,
                p0_coverage, p1_coverage, p2_coverage, p3_coverage,
                tc_pass_count, tc_total, cp_covered, cp_total
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            project_id,
            today,
            coverage_data['actual_coverage'],
            coverage_data.get('p0_coverage', 0),
            coverage_data.get('p1_coverage', 0),
            coverage_data.get('p2_coverage', 0),
            coverage_data.get('p3_coverage', 0),
            coverage_data['tc_pass_count'],
            coverage_data['tc_total'],
            coverage_data['cp_covered'],
            coverage_data['cp_total']
        ))
        conn.commit()
        snapshot_id = cursor.lastrowid
    
    # 获取创建的快照
    cursor.execute("SELECT * FROM project_progress WHERE id = ?", (snapshot_id,))
    row = cursor.fetchone()
    snapshot = {
        'id': row[0],
        'snapshot_date': row[2],
        'actual_coverage': row[3],
        'tc_pass_count': row[4],
        'tc_total': row[5],
        'cp_covered': row[6],
        'cp_total': row[7],
        'created_at': row[8]
    }
    
    return jsonify({
        "success": True,
        "snapshot": snapshot
    })


@api.route("/api/progress/<int:project_id>/snapshots", methods=["GET"])
@login_required
def get_snapshots(project_id):
    """获取项目快照列表"""
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    
    if not project:
        return jsonify({"error": "项目不存在"}), 404
    
    # 确保表存在
    ensure_progress_table_exists(project["name"])
    conn = get_db(project["name"])
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT id, snapshot_date, actual_coverage, tc_pass_count, tc_total, 
               cp_covered, cp_total, created_at
        FROM project_progress 
        WHERE project_id = ?
        ORDER BY snapshot_date DESC
    """, (project_id,))
    
    snapshots = []
    for row in cursor.fetchall():
        snapshots.append({
            'id': row[0],
            'snapshot_date': row[1],
            'actual_coverage': row[2],
            'tc_pass_count': row[3],
            'tc_total': row[4],
            'cp_covered': row[5],
            'cp_total': row[6],
            'created_at': row[7]
        })
    
    return jsonify({
        "snapshots": snapshots
    })


@api.route("/api/progress/snapshots/<int:snapshot_id>", methods=["DELETE"])
@admin_required
def delete_snapshot(snapshot_id):
    """删除快照"""
    from datetime import datetime
    
    # 先获取快照信息
    conn = None
    for project in load_projects():
        try:
            temp_conn = get_db(project["name"])
            temp_cursor = temp_conn.cursor()
            temp_cursor.execute("SELECT project_id, snapshot_date FROM project_progress WHERE id = ?", (snapshot_id,))
            row = temp_cursor.fetchone()
            if row:
                conn = temp_conn
                project_id = row[0]
                snapshot_date = row[1]
                break
        except Exception:
            continue
    
    if not conn:
        return jsonify({"error": "快照不存在"}), 404
    
    # 检查是否当周快照
    today = datetime.now().strftime('%Y-%m-%d')
    is_current_week = snapshot_date == today
    
    # 删除快照
    cursor = conn.cursor()
    cursor.execute("DELETE FROM project_progress WHERE id = ?", (snapshot_id,))
    conn.commit()
    
    return jsonify({
        "success": True,
        "message": "Snapshot deleted",
        "is_current_week": is_current_week
    })


@api.route("/api/progress/<int:project_id>/export", methods=["GET"])
@login_required
def export_progress(project_id):
    """导出项目进度数据"""
    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)
    
    if not project:
        return jsonify({"error": "项目不存在"}), 404
    
    # 确保表存在
    ensure_progress_table_exists(project["name"])
    conn = get_db(project["name"])
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT snapshot_date, actual_coverage, tc_pass_count, tc_total, 
               cp_covered, cp_total, created_at
        FROM project_progress 
        WHERE project_id = ?
        ORDER BY snapshot_date
    """, (project_id,))
    
    # CSV 格式导出
    csv_lines = ["snapshot_date,actual_coverage,tc_pass_count,tc_total,cp_covered,cp_total,created_at"]
    for row in cursor.fetchall():
        csv_lines.append(f"{row[0]},{row[1]},{row[2]},{row[3]},{row[4]},{row[5]},{row[6]}")
    
    return "\n".join(csv_lines), 200, {
        'Content-Type': 'text/csv',
        'Content-Disposition': f'attachment; filename=progress_{project["name"]}.csv'
    }


# v0.8.2: 定时任务接口（需要 API Token）
@api.route("/api/cron/progress-snapshot", methods=["POST"])
def cron_progress_snapshot():
    """定时任务：批量创建所有项目的快照"""
    # 验证 API Token
    token = request.headers.get('X-API-Token')
    expected_token = current_app.config.get('CRON_API_TOKEN')
    
    if not token or token != expected_token:
        return jsonify({"error": "Unauthorized"}), 401
    
    if not expected_token:
        return jsonify({"error": "Cron not configured"}), 500
    
    projects = load_projects()
    created_count = 0
    today = datetime.now().strftime('%Y-%m-%d')
    
    for project in projects:
        if project.get('is_archived'):
            continue
        
        project_name = project["name"]
        project_id = project["id"]
        
        # 检查项目是否有日期
        if not project.get('start_date') or not project.get('end_date'):
            continue
        
        # 确保表存在
        try:
            ensure_progress_table_exists(project_name)
        except Exception:
            continue
        
        # 计算覆盖率
        coverage_data = calculate_current_coverage(project_name)
        if coverage_data is None:
            continue
        
        # 检查是否已存在当周快照
        conn = get_db(project_name)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id FROM project_progress 
            WHERE project_id = ? AND snapshot_date = ?
        """, (project_id, today))
        
        if not cursor.fetchone():
            cursor.execute("""
                INSERT INTO project_progress (
                    project_id, snapshot_date, actual_coverage,
                    tc_pass_count, tc_total, cp_covered, cp_total
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                project_id,
                today,
                coverage_data['actual_coverage'],
                coverage_data['tc_pass_count'],
                coverage_data['tc_total'],
                coverage_data['cp_covered'],
                coverage_data['cp_total']
            ))
            conn.commit()
            created_count += 1
    
    return jsonify({
        "success": True,
        "message": "Snapshots created",
        "count": created_count
    })


@api.route("/api/cp", methods=["GET"])
def get_coverpoints():
    """获取 CP 列表（含覆盖率计算和过滤）"""
    project_id = request.args.get("project_id", type=int)
    feature_filter = request.args.get("feature")
    priority_filter = request.args.get("priority")
    # REQ-005: 支持 filter 参数 (all/unlinked)
    link_filter = request.args.get("filter")

    if not project_id:
        return jsonify([])

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify([])

    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 构建过滤查询
    query = "SELECT * FROM cover_point WHERE 1=1"
    params = []

    if feature_filter:
        features = [f.strip() for f in feature_filter.split(",")]
        placeholders = ",".join(["?"] * len(features))
        query += f" AND feature IN ({placeholders})"
        params.extend(features)

    if priority_filter:
        priorities = [p.strip() for p in priority_filter.split(",")]
        placeholders = ",".join(["?"] * len(priorities))
        query += f" AND priority IN ({placeholders})"
        params.extend(priorities)

    query += " ORDER BY id"
    cursor.execute(query, params)

    # REQ-005: 如果需要过滤未关联的CP，先获取所有已关联的CP ID
    linked_cp_ids = set()
    if link_filter == "unlinked":
        cursor.execute("SELECT DISTINCT cp_id FROM tc_cp_connections")
        linked_cp_ids = {row["cp_id"] for row in cursor.fetchall()}
        # 重新执行原始查询
        cursor.execute(query, params)

    cps = []
    for row in cursor.fetchall():
        cp_id = row["id"]

        # 计算覆盖率：统计关联 TC 中 PASS 的比例
        cursor.execute(
            """
            SELECT tc.status FROM test_case tc
            INNER JOIN tc_cp_connections tcc ON tc.id = tcc.tc_id
            WHERE tcc.cp_id = ?
        """,
            (cp_id,),
        )

        connected_tcs = cursor.fetchall()
        total = len(connected_tcs)
        passed = sum(1 for tc in connected_tcs if tc["status"] == "PASS")

        # 计算覆盖率
        coverage = round(passed / total * 100, 1) if total > 0 else 0.0

        # REQ-005: 判断是否关联
        is_linked = cp_id in linked_cp_ids or (total > 0)
        if link_filter == "unlinked" and is_linked:
            continue  # 跳过已关联的CP

        cps.append(
            {
                "id": row["id"],
                "project_id": row["project_id"],
                "feature": row["feature"],
                "sub_feature": row["sub_feature"],
                "cover_point": row["cover_point"],
                "cover_point_details": row["cover_point_details"],
                "comments": row["comments"],
                "priority": row["priority"],
                "created_at": row["created_at"],
                "created_by": dict(row).get("created_by", ""),
                "coverage": coverage,
                "coverage_detail": f"{passed}/{total}",
                "linked": is_linked,
            }
        )

    return jsonify(cps)


@api.route("/api/cp", methods=["POST"])
@guest_required
def create_coverpoint():
    """创建 CP"""
    data = request.json
    project_id = data.get("project_id")

    if not project_id:
        return jsonify({"error": "需要指定项目"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    # 获取当前登录用户
    current_username = session.get(SESSION_USERNAME_KEY, "anonymous")

    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 确保 created_by 列存在
    try:
        cursor.execute("ALTER TABLE cover_point ADD COLUMN created_by TEXT")
    except sqlite3.OperationalError:
        pass  # 列已存在

    today = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor.execute(
        """
        INSERT INTO cover_point (project_id, feature, sub_feature, cover_point, cover_point_details, comments, priority, created_at, created_by)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """,
        (
            project_id,
            data.get("feature", ""),
            data.get("sub_feature", ""),
            data.get("cover_point", ""),
            data.get("cover_point_details", ""),
            data.get("comments", ""),
            data.get("priority", "P0"),
            today,
            current_username,
        ),
    )

    conn.commit()

    cp_id = cursor.lastrowid
    return jsonify(
        {
            "success": True,
            "item": {
                "id": cp_id,
                "project_id": project_id,
                "feature": data.get("feature", ""),
                "sub_feature": data.get("sub_feature", ""),
                "cover_point": data.get("cover_point", ""),
                "cover_point_details": data.get("cover_point_details", ""),
                "comments": data.get("comments", ""),
                "priority": data.get("priority", "P0"),
                "created_at": today,
                "created_by": current_username,
            },
        }
    )


@api.route("/api/cp/<int:cp_id>", methods=["GET"])
def get_coverpoint(cp_id):
    """获取 CP 详情"""
    project_id = request.args.get("project_id", type=int)

    if not project_id:
        return jsonify({"error": "需要指定项目"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM cover_point WHERE id = ?", (cp_id,))
    cp = cursor.fetchone()

    if not cp:
        return jsonify({"error": "Cover Point 不存在"}), 404

    # 将 sqlite3.Row 转换为字典
    cp_dict = dict(cp)

    # 处理 created_by 字段（可能不存在于旧数据中）
    created_by = cp_dict.get("created_by", "") or ""

    return jsonify(
        {
            "id": cp_dict["id"],
            "project_id": cp_dict["project_id"],
            "feature": cp_dict["feature"],
            "sub_feature": cp_dict["sub_feature"],
            "cover_point": cp_dict["cover_point"],
            "cover_point_details": cp_dict["cover_point_details"],
            "comments": cp_dict["comments"],
            "priority": cp_dict["priority"],
            "created_at": cp_dict["created_at"],
            "created_by": created_by,
        }
    )


@api.route("/api/cp/<int:cp_id>", methods=["PUT"])
@guest_required
def update_coverpoint(cp_id):
    """更新 CP"""
    data = request.json
    project_id = data.get("project_id")

    if not project_id:
        return jsonify({"error": "需要指定项目"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 获取当前 CP 的 priority 值（如果请求体中没有提供）
    cursor.execute("SELECT priority FROM cover_point WHERE id=?", (cp_id,))
    current = cursor.fetchone()
    current_priority = current["priority"] if current else "P0"

    # 如果请求体中没有 priority，保留当前值
    new_priority = data.get("priority", current_priority)

    cursor.execute(
        """
        UPDATE cover_point SET feature=?, sub_feature=?, cover_point=?, cover_point_details=?, comments=?, priority=?
        WHERE id=?
    """,
        (
            data.get("feature", ""),
            data.get("sub_feature", ""),
            data.get("cover_point", ""),
            data.get("cover_point_details", ""),
            data.get("comments", ""),
            new_priority,
            cp_id,
        ),
    )

    conn.commit()

    return jsonify({"success": True})


@api.route("/api/cp/<int:cp_id>", methods=["DELETE"])
@guest_required
def delete_coverpoint(cp_id):
    """删除 CP"""
    project_id = request.args.get("project_id", type=int)

    if not project_id:
        return jsonify({"error": "需要指定项目"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    cursor.execute("DELETE FROM cover_point WHERE id=?", (cp_id,))
    cursor.execute("DELETE FROM tc_cp_connections WHERE cp_id=?", (cp_id,))

    conn.commit()

    return jsonify({"success": True})


@api.route("/api/cp/<int:cp_id>/tcs", methods=["GET"])
def get_cp_tcs(cp_id):
    """获取 CP 关联的 TC 列表"""
    project_id = request.args.get("project_id", type=int)
    projects = load_projects()

    # 如果提供了 project_id，直接使用
    project = None
    if project_id:
        project = next((p for p in projects if p["id"] == project_id), None)

    # 如果没找到，遍历所有项目查找
    if not project:
        for p in projects:
            conn = get_db(p["name"])
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT id FROM cover_point WHERE id=?", (cp_id,))
                if cursor.fetchone():
                    project = p
                    break
            except sqlite3.OperationalError:
                continue

    if not project:
        return jsonify({"error": "CP 不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 获取关联的 TC
    try:
        cursor.execute(
            """
            SELECT tc.id, tc.test_name, tc.status
            FROM test_case tc
            INNER JOIN tc_cp_connections tcc ON tc.id = tcc.tc_id
            WHERE tcc.cp_id = ?
            ORDER BY tc.id
        """,
            (cp_id,),
        )
    except sqlite3.OperationalError:
        return jsonify({"cp_id": cp_id, "connected_tcs": []})

    tcs = []
    for row in cursor.fetchall():
        tcs.append({"id": row["id"], "test_name": row["test_name"], "status": row["status"]})

    return jsonify({"cp_id": cp_id, "connected_tcs": tcs})


# ============ Test Cases ============


@api.route("/api/tc", methods=["GET"])
def get_testcases():
    """获取 TC 列表（支持过滤，v0.6.2）"""
    project_id = request.args.get("project_id", type=int)
    sort_by = request.args.get("sort_by", "id")
    status_filter = request.args.get("status")
    dv_milestone_filter = request.args.get("dv_milestone")
    priority_filter = request.args.get("priority")
    owner_filter = request.args.get("owner")
    category_filter = request.args.get("category")
    search = request.args.get("search", "")

    if not project_id:
        return jsonify([])

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify([])

    conn = get_db(project["name"])
    cursor = conn.cursor()

    query = "SELECT * FROM test_case WHERE 1=1"
    params = []

    # Status 过滤（支持多值，逗号分隔）
    if status_filter:
        statuses = [s.strip() for s in status_filter.split(",")]
        placeholders = ",".join(["?"] * len(statuses))
        query += f" AND status IN ({placeholders})"
        params.extend(statuses)
        placeholders = ",".join(["?"] * len(statuses))
        query += f" AND status IN ({placeholders})"
        params.extend(statuses)

    # DV Milestone 过滤（支持多值，逗号分隔）
    if dv_milestone_filter:
        milestones = [m.strip() for m in dv_milestone_filter.split(",")]
        placeholders = ",".join(["?"] * len(milestones))
        query += f" AND dv_milestone IN ({placeholders})"
        params.extend(milestones)

    # Priority 过滤
    if priority_filter:
        priorities = [p.strip() for p in priority_filter.split(",")]
        placeholders = ",".join(["?"] * len(priorities))
        query += f" AND priority IN ({placeholders})"
        params.extend(priorities)

    # Owner 过滤（支持多值，逗号分隔）
    if owner_filter:
        owners = [o.strip() for o in owner_filter.split(",")]
        placeholders = ",".join(["?"] * len(owners))
        query += f" AND owner IN ({placeholders})"
        params.extend(owners)

    # Category 过滤（支持多值，逗号分隔）
    if category_filter:
        categories = [c.strip() for c in category_filter.split(",")]
        placeholders = ",".join(["?"] * len(categories))
        query += f" AND category IN ({placeholders})"
        params.extend(categories)

    # 保留原有 status_filter 参数（兼容旧版本）
    if status_filter and not status_filter:
        query += " AND status=?"
        params.append(status_filter)

    if search:
        query += " AND (test_name LIKE ? OR testbench LIKE ? OR owner LIKE ?)"
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])

    # 排序
    if sort_by == "testbench":
        query += " ORDER BY testbench"
    elif sort_by == "owner":
        query += " ORDER BY owner"
    elif sort_by == "status":
        query += " ORDER BY status"
    elif sort_by == "created_at":
        query += " ORDER BY created_at"
    else:
        query += " ORDER BY id"

    cursor.execute(query, params)

    tcs = []
    for row in cursor.fetchall():
        # 获取关联的 CP
        cursor.execute("SELECT cp_id FROM tc_cp_connections WHERE tc_id=?", (row["id"],))
        connected_cps = [r[0] for r in cursor.fetchall()]

        tcs.append(
            {
                "id": row["id"],
                "project_id": row["project_id"],
                "dv_milestone": row["dv_milestone"],
                "testbench": row["testbench"],
                "category": row["category"],
                "owner": row["owner"],
                "test_name": row["test_name"],
                "scenario_details": row["scenario_details"],
                "checker_details": row["checker_details"],
                "coverage_details": row["coverage_details"],
                "comments": row["comments"],
                "status": row["status"],
                "created_at": row["created_at"],
                "coded_date": row["coded_date"],
                "fail_date": row["fail_date"],
                "pass_date": row["pass_date"],
                "removed_date": row["removed_date"],
                "target_date": row["target_date"],
                "created_by": dict(row).get("created_by", ""),
                "connected_cps": connected_cps,
            }
        )

    return jsonify(tcs)


@api.route("/api/tc/<int:tc_id>", methods=["GET"])
def get_testcase(tc_id):
    """获取 TC 详情"""
    project_id = request.args.get("project_id", type=int)

    if not project_id:
        return jsonify({"error": "需要指定项目"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM test_case WHERE id = ?", (tc_id,))
    tc = cursor.fetchone()

    if not tc:
        return jsonify({"error": "Test Case 不存在"}), 404

    # 将 sqlite3.Row 转换为字典
    tc_dict = dict(tc)

    # 获取关联的 CP
    cursor.execute(
        """
        SELECT cp.id, cp.cover_point FROM cover_point cp
        INNER JOIN tc_cp_connections tcc ON cp.id = tcc.cp_id
        WHERE tcc.tc_id = ?
    """,
        (tc_id,),
    )
    connected_cps = [row["id"] for row in cursor.fetchall()]

    # 处理 created_by 字段（可能不存在于旧数据中）
    created_by = tc_dict.get("created_by", "") or ""

    return jsonify(
        {
            "id": tc_dict["id"],
            "project_id": tc_dict["project_id"],
            "dv_milestone": tc_dict["dv_milestone"],
            "testbench": tc_dict["testbench"],
            "category": tc_dict["category"],
            "owner": tc_dict["owner"],
            "test_name": tc_dict["test_name"],
            "scenario_details": tc_dict["scenario_details"],
            "checker_details": tc_dict["checker_details"],
            "coverage_details": tc_dict["coverage_details"],
            "comments": tc_dict["comments"],
            "status": tc_dict["status"],
            "created_at": tc_dict["created_at"],
            "coded_date": tc_dict["coded_date"],
            "fail_date": tc_dict["fail_date"],
            "pass_date": tc_dict["pass_date"],
            "removed_date": tc_dict["removed_date"],
            "target_date": tc_dict["target_date"],
            "connected_cps": connected_cps,
            "created_by": created_by,
        }
    )


@api.route("/api/tc", methods=["POST"])
@guest_required
def create_testcase():
    """创建 TC"""
    data = request.json
    project_id = data.get("project_id")

    if not project_id:
        return jsonify({"error": "需要指定项目"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    # 获取当前登录用户
    current_username = session.get(SESSION_USERNAME_KEY, "anonymous")

    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 确保 created_by 列存在
    try:
        cursor.execute("ALTER TABLE test_case ADD COLUMN created_by TEXT")
    except sqlite3.OperationalError:
        pass  # 列已存在

    today = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor.execute(
        """
        INSERT INTO test_case (
            project_id, dv_milestone, testbench, category, owner, test_name, 
            scenario_details, checker_details, coverage_details, comments, priority,
            status, created_at, coded_date, fail_date, pass_date, removed_date, target_date, created_by
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'OPEN', ?, NULL, NULL, NULL, NULL, ?, ?)
    """,
        (
            project_id,
            data.get("dv_milestone", ""),
            data.get("testbench", ""),
            data.get("category", ""),
            data.get("owner", ""),
            data.get("test_name", ""),
            data.get("scenario_details", ""),
            data.get("checker_details", ""),
            data.get("coverage_details", ""),
            data.get("comments", ""),
            data.get("priority", "P0"),
            today,
            data.get("target_date", ""),
            current_username,
        ),
    )

    tc_id = cursor.lastrowid

    # 关联 CP
    if data.get("connections"):
        for cp_id in data["connections"]:
            try:
                cursor.execute(
                    "INSERT INTO tc_cp_connections (tc_id, cp_id) VALUES (?, ?)", (tc_id, cp_id)
                )
            except sqlite3.IntegrityError:
                pass

    conn.commit()

    return jsonify(
        {
            "success": True,
            "item": {
                "id": tc_id,
                "project_id": project_id,
                "dv_milestone": data.get("dv_milestone", ""),
                "testbench": data.get("testbench", ""),
                "category": data.get("category", ""),
                "owner": data.get("owner", ""),
                "test_name": data.get("test_name", ""),
                "scenario_details": data.get("scenario_details", ""),
                "checker_details": data.get("checker_details", ""),
                "coverage_details": data.get("coverage_details", ""),
                "comments": data.get("comments", ""),
                "priority": data.get("priority", "P0"),
                "status": "OPEN",
                "created_at": today,
                "coded_date": None,
                "fail_date": None,
                "pass_date": None,
                "removed_date": None,
                "target_date": data.get("target_date", ""),
                "connected_cps": data.get("connections", []),
                "created_by": current_username,
            },
        }
    )


@api.route("/api/tc/<int:tc_id>", methods=["PUT"])
@guest_required
def update_testcase(tc_id):
    """更新 TC"""
    data = request.json
    project_id = data.get("project_id")

    if not project_id:
        return jsonify({"error": "需要指定项目"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    cursor.execute(
        """
        UPDATE test_case SET 
            dv_milestone=?, 
            testbench=?, 
            category=?, 
            owner=?, 
            test_name=?, 
            scenario_details=?, 
            checker_details=?, 
            coverage_details=?, 
            comments=?,
            target_date=?
        WHERE id=?
    """,
        (
            data.get("dv_milestone", ""),
            data.get("testbench", ""),
            data.get("category", ""),
            data.get("owner", ""),
            data.get("test_name", ""),
            data.get("scenario_details", ""),
            data.get("checker_details", ""),
            data.get("coverage_details", ""),
            data.get("comments", ""),
            data.get("target_date", ""),
            tc_id,
        ),
    )

    # 更新关联
    if "connections" in data:
        cursor.execute("DELETE FROM tc_cp_connections WHERE tc_id=?", (tc_id,))
        for cp_id in data["connections"]:
            try:
                cursor.execute(
                    "INSERT INTO tc_cp_connections (tc_id, cp_id) VALUES (?, ?)", (tc_id, cp_id)
                )
            except sqlite3.IntegrityError:
                pass

    conn.commit()

    return jsonify({"success": True})


@api.route("/api/tc/<int:tc_id>", methods=["DELETE"])
@guest_required
def delete_testcase(tc_id):
    """删除 TC"""
    project_id = request.args.get("project_id", type=int)

    if not project_id:
        return jsonify({"error": "需要指定项目"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    cursor.execute("DELETE FROM test_case WHERE id=?", (tc_id,))
    cursor.execute("DELETE FROM tc_cp_connections WHERE tc_id=?", (tc_id,))

    conn.commit()

    return jsonify({"success": True})


@api.route("/api/tc/<int:tc_id>/status", methods=["POST"])
def update_status(tc_id):
    """更新状态"""
    data = request.json
    new_status = data.get("status")
    project_id = data.get("project_id")

    if not project_id:
        return jsonify({"error": "需要指定项目"}), 400

    valid_statuses = ["OPEN", "CODED", "FAIL", "PASS", "REMOVED"]
    if new_status not in valid_statuses:
        return jsonify({"error": "无效状态"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 获取当前 TC 信息
    cursor.execute("SELECT status FROM test_case WHERE id=?", (tc_id,))
    row = cursor.fetchone()
    if not row:
        return jsonify({"error": "TC 不存在"}), 404

    old_status = row["status"]

    # 状态日期映射
    status_dates = {
        "CODED": "coded_date",
        "FAIL": "fail_date",
        "PASS": "pass_date",
        "REMOVED": "removed_date",
    }

    today = datetime.now().strftime("%Y-%m-%d")

    # 清除所有状态日期
    cursor.execute(
        """
        UPDATE test_case SET 
            coded_date=NULL, 
            fail_date=NULL, 
            pass_date=NULL, 
            removed_date=NULL 
        WHERE id=?
    """,
        (tc_id,),
    )

    # 设置新状态对应的日期
    if new_status in status_dates:
        date_field = status_dates[new_status]
        cursor.execute(f"UPDATE test_case SET {date_field}=? WHERE id=?", (today, tc_id))

    # 如果是 REMOVED，清除 CP 关联
    if new_status == "REMOVED":
        cursor.execute("DELETE FROM tc_cp_connections WHERE tc_id=?", (tc_id,))

    # 更新状态
    cursor.execute("UPDATE test_case SET status=? WHERE id=?", (new_status, tc_id))

    conn.commit()

    # 检查是否需要确认（从 PASS 改为其他状态）
    need_confirm = old_status == "PASS" and new_status != "PASS"

    return jsonify({"success": True, "status": new_status, "need_confirm": need_confirm})


# ============ 批量操作 ============


@api.route("/api/tc/batch/status", methods=["POST"])
@guest_required
def batch_update_tc_status():
    """批量更新 TC 状态"""
    data = request.json
    project_id = data.get("project_id")
    tc_ids = data.get("tc_ids", [])
    new_status = data.get("status")

    if not project_id or not tc_ids or not new_status:
        return jsonify({"error": "缺少必要参数"}), 400

    valid_statuses = ["OPEN", "CODED", "FAIL", "PASS", "REMOVED"]
    if new_status not in valid_statuses:
        return jsonify({"error": "无效状态"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 状态日期映射
    status_dates = {
        "CODED": "coded_date",
        "FAIL": "fail_date",
        "PASS": "pass_date",
        "REMOVED": "removed_date",
    }

    today = datetime.now().strftime("%Y-%m-%d")

    success_count = 0
    for tc_id in tc_ids:
        # 检查 TC 是否存在
        cursor.execute("SELECT status FROM test_case WHERE id=?", (tc_id,))
        row = cursor.fetchone()
        if not row:
            continue

        # 清除所有状态日期
        cursor.execute(
            """
            UPDATE test_case SET 
                coded_date=NULL, 
                fail_date=NULL, 
                pass_date=NULL, 
                removed_date=NULL 
            WHERE id=?
        """,
            (tc_id,),
        )

        # 设置新状态对应的日期
        if new_status in status_dates:
            date_field = status_dates[new_status]
            cursor.execute(f"UPDATE test_case SET {date_field}=? WHERE id=?", (today, tc_id))

        # 如果是 REMOVED，清除 CP 关联
        if new_status == "REMOVED":
            cursor.execute("DELETE FROM tc_cp_connections WHERE tc_id=?", (tc_id,))

        # 更新状态
        cursor.execute("UPDATE test_case SET status=? WHERE id=?", (new_status, tc_id))
        success_count += 1

    conn.commit()

    return jsonify({"success": success_count, "failed": len(tc_ids) - success_count})

@guest_required

@api.route("/api/tc/batch/target_date", methods=["POST"])
@guest_required
def batch_update_tc_target_date():
    """批量更新 TC Target Date"""
    data = request.json
    project_id = data.get("project_id")
    tc_ids = data.get("tc_ids", [])
    target_date = data.get("target_date")

    if not project_id or not tc_ids:
        return jsonify({"error": "缺少必要参数"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    success_count = 0
    for tc_id in tc_ids:
        cursor.execute("UPDATE test_case SET target_date=? WHERE id=?", (target_date, tc_id))
        success_count += 1

    conn.commit()

    return jsonify({"success": success_count, "failed": len(tc_ids) - success_count})
@guest_required

@guest_required

@api.route("/api/tc/batch/dv_milestone", methods=["POST"])
@guest_required
def batch_update_tc_dv_milestone():
    """批量更新 TC DV Milestone"""
    data = request.json
    project_id = data.get("project_id")
    tc_ids = data.get("tc_ids", [])
    dv_milestone = data.get("dv_milestone")

    if not project_id or not tc_ids:
        return jsonify({"error": "缺少必要参数"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    success_count = 0
    for tc_id in tc_ids:
        cursor.execute("UPDATE test_case SET dv_milestone=? WHERE id=?", (dv_milestone, tc_id))
        success_count += 1

    conn.commit()

    return jsonify({"success": success_count, "failed": len(tc_ids) - success_count})


@api.route("/api/cp/batch/priority", methods=["POST"])
@guest_required
def batch_update_cp_priority():
    """批量更新 CP Priority"""
    data = request.json
    project_id = data.get("project_id")
    cp_ids = data.get("cp_ids", [])
    priority = data.get("priority")

    if not project_id or not cp_ids:
        return jsonify({"error": "缺少必要参数"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    success_count = 0
    for cp_id in cp_ids:
        cursor.execute("UPDATE cover_point SET priority=? WHERE id=?", (priority, cp_id))
        success_count += 1

    conn.commit()

    return jsonify({"success": success_count, "failed": len(cp_ids) - success_count})


# ============ Functional Coverage (FC) API - v0.11.0 ============


@api.route("/api/fc", methods=["GET"])
@login_required
def get_fc_list():
    """获取 FC 列表（支持筛选）"""
    project_id = request.args.get("project_id", type=int)
    covergroup_filter = request.args.get("covergroup")
    coverpoint_filter = request.args.get("coverpoint")
    coverage_type_filter = request.args.get("coverage_type")
    bin_name_search = request.args.get("bin_name")  # 模糊搜索

    if not project_id:
        return jsonify([])

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify([])

    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 构建过滤查询
    query = "SELECT * FROM functional_coverage WHERE 1=1"
    params = []

    if covergroup_filter:
        groups = [g.strip() for g in covergroup_filter.split(",")]
        placeholders = ",".join(["?"] * len(groups))
        query += f" AND covergroup IN ({placeholders})"
        params.extend(groups)

    if coverpoint_filter:
        points = [p.strip() for p in coverpoint_filter.split(",")]
        placeholders = ",".join(["?"] * len(points))
        query += f" AND coverpoint IN ({placeholders})"
        params.extend(points)

    if coverage_type_filter:
        types = [t.strip() for t in coverage_type_filter.split(",")]
        placeholders = ",".join(["?"] * len(types))
        query += f" AND coverage_type IN ({placeholders})"
        params.extend(types)

    if bin_name_search:
        query += " AND bin_name LIKE ?"
        params.append(f"%{bin_name_search}%")

    query += " ORDER BY covergroup, coverpoint, bin_name"
    cursor.execute(query, params)

    fc_list = [dict(row) for row in cursor.fetchall()]
    return jsonify(fc_list)


@api.route("/api/fc/import", methods=["POST"])
@login_required
def import_fc():
    """导入 FC (CSV)"""
    project_id = request.args.get("project_id", type=int)

    if not project_id:
        return jsonify({"error": "缺少 project_id"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    # 获取 CSV 数据
    csv_data = request.json.get("csv_data", [])
    if not csv_data or len(csv_data) < 2:
        return jsonify({"error": "CSV 数据为空或格式错误"}), 400

    headers = csv_data[0]
    header_map = {}
    for idx, header in enumerate(headers):
        if header:
            header_map[header.strip()] = idx

    # 检查必填字段
    required_fields = ["Covergroup", "Coverpoint", "Type", "Bin_Name"]
    for field in required_fields:
        if field not in header_map:
            return jsonify({"error": f"缺少必填字段: {field}"}), 400

    conn = get_db(project["name"])
    cursor = conn.cursor()

    imported = 0
    errors = []
    created_by = session.get(SESSION_USERNAME_KEY, "unknown")

    for row_idx, row in enumerate(csv_data[1:], start=2):
        try:
            covergroup = (
                row[header_map.get("Covergroup", 0)]
                if header_map.get("Covergroup", 0) < len(row)
                else None
            )
            coverpoint = (
                row[header_map.get("Coverpoint", 1)]
                if header_map.get("Coverpoint", 1) < len(row)
                else None
            )
            coverage_type = (
                row[header_map.get("Type", 2)]
                if header_map.get("Type", 2) < len(row)
                else None
            )
            bin_name = (
                row[header_map.get("Bin_Name", 3)]
                if header_map.get("Bin_Name", 3) < len(row)
                else None
            )

            if not all([covergroup, coverpoint, coverage_type, bin_name]):
                errors.append(f"第{row_idx}行: 必填字段缺失")
                continue

            # 检查重名 (project_id + covergroup + coverpoint + bin_name)
            cursor.execute(
                "SELECT id FROM functional_coverage WHERE project_id=? AND covergroup=? AND coverpoint=? AND bin_name=?",
                (project["id"], covergroup, coverpoint, bin_name)
            )
            if cursor.fetchone():
                errors.append(f'第{row_idx}行: FC "{covergroup}/{coverpoint}/{bin_name}" 已存在')
                continue

            # 获取可选字段
            bin_val = (
                row[header_map.get("Bin_Value", 4)]
                if header_map.get("Bin_Value", 4) < len(row)
                else None
            )
            coverage_pct = 0.0
            if header_map.get("Coverage_Pct", 5) < len(row):
                try:
                    coverage_pct = float(row[header_map.get("Coverage_Pct", 5)] or 0)
                except (ValueError, TypeError):
                    coverage_pct = 0.0
            status = (
                row[header_map.get("Status", 6)]
                if header_map.get("Status", 6) < len(row)
                else "missing"
            )
            # 空字符串或无效值默认为 'missing'
            if not status or status.strip() == '':
                status = "missing"
            # 校验 status 只能是 'missing' 或 'ready'
            if status not in ('missing', 'ready'):
                errors.append(f'第{row_idx}行: Status 必须是 "missing" 或 "ready"，当前值为 "{status}"')
                continue
            comments = (
                row[header_map.get("Comments", 7)]
                if header_map.get("Comments", 7) < len(row)
                else None
            )

            cursor.execute(
                """
                INSERT INTO functional_coverage (project_id, covergroup, coverpoint, coverage_type, bin_name, bin_val, coverage_pct, status, comments, created_by, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))
                """,
                (project["id"], covergroup, coverpoint, coverage_type, bin_name, bin_val, coverage_pct, status, comments, created_by)
            )

            imported += 1
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")

    if imported > 0:
        conn.commit()

    return jsonify({"success": True, "imported": imported, "failed": len(errors), "errors": errors})


@api.route("/api/fc/export", methods=["GET"])
@login_required
def export_fc():
    """导出 FC (CSV)"""
    project_id = request.args.get("project_id", type=int)

    if not project_id:
        return jsonify({"error": "缺少 project_id"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM functional_coverage ORDER BY covergroup, coverpoint, bin_name")
    fc_list = [dict(row) for row in cursor.fetchall()]

    # 生成 CSV
    headers = ["Covergroup", "Coverpoint", "Type", "Bin_Name", "Bin_Value", "Coverage_Pct", "Status", "Comments"]
    csv_data = [headers]

    for fc in fc_list:
        csv_data.append([
            fc.get("covergroup", ""),
            fc.get("coverpoint", ""),
            fc.get("coverage_type", ""),
            fc.get("bin_name", ""),
            fc.get("bin_val", ""),
            str(fc.get("coverage_pct", 0.0)),
            fc.get("status", "missing"),
            fc.get("comments", "") or ""
        ])

    return jsonify({"success": True, "csv_data": csv_data})


# ============ FC-CP 关联 API - v0.11.0 ============


@api.route("/api/fc-cp-association", methods=["GET"])
@login_required
def get_fc_cp_associations():
    """获取 FC-CP 关联列表"""
    project_id = request.args.get("project_id", type=int)

    if not project_id:
        return jsonify([])

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify([])

    conn = get_db(project["name"])
    cursor = conn.cursor()

    cursor.execute("""
        SELECT fcca.*,
               cp.feature as cp_feature, cp.sub_feature as cp_sub_feature, cp.cover_point as cp_cover_point,
               fc.covergroup as fc_covergroup, fc.coverpoint as fc_coverpoint, fc.bin_name as fc_bin_name
        FROM fc_cp_association fcca
        LEFT JOIN cover_point cp ON fcca.cp_id = cp.id
        LEFT JOIN functional_coverage fc ON fcca.fc_id = fc.id
        ORDER BY cp.feature, cp.sub_feature, cp.cover_point
    """)

    associations = []
    for row in cursor.fetchall():
        assoc = dict(row)
        # 添加 CP 和 FC 的详细信息
        associations.append({
            "id": assoc["id"],
            "project_id": assoc["project_id"],
            "cp_id": assoc["cp_id"],
            "fc_id": assoc["fc_id"],
            "created_by": assoc.get("created_by"),
            "created_at": assoc.get("created_at"),
            "cp_feature": assoc.get("cp_feature"),
            "cp_sub_feature": assoc.get("cp_sub_feature"),
            "cp_cover_point": assoc.get("cp_cover_point"),
            "fc_covergroup": assoc.get("fc_covergroup"),
            "fc_coverpoint": assoc.get("fc_coverpoint"),
            "fc_bin_name": assoc.get("fc_bin_name")
        })

    return jsonify(associations)


@api.route("/api/fc-cp-association", methods=["POST"])
@login_required
def create_fc_cp_association():
    """创建 FC-CP 关联"""
    project_id = request.json.get("project_id")
    cp_id = request.json.get("cp_id")
    fc_id = request.json.get("fc_id")

    if not all([project_id, cp_id, fc_id]):
        return jsonify({"error": "缺少必要参数"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 检查 CP 是否存在
    cursor.execute("SELECT id FROM cover_point WHERE id=?", (cp_id,))
    if not cursor.fetchone():
        return jsonify({"error": "CP 不存在"}), 404

    # 检查 FC 是否存在
    cursor.execute("SELECT id FROM functional_coverage WHERE id=?", (fc_id,))
    if not cursor.fetchone():
        return jsonify({"error": "FC 不存在"}), 404

    # 检查关联是否已存在
    cursor.execute("SELECT id FROM fc_cp_association WHERE cp_id=? AND fc_id=?", (cp_id, fc_id))
    if cursor.fetchone():
        return jsonify({"error": "关联已存在"}), 400

    created_by = session.get(SESSION_USERNAME_KEY, "unknown")

    try:
        cursor.execute(
            """
            INSERT INTO fc_cp_association (project_id, cp_id, fc_id, created_by, created_at)
            VALUES (?, ?, ?, ?, datetime('now'))
            """,
            (project_id, cp_id, fc_id, created_by)
        )
        conn.commit()
        assoc_id = cursor.lastrowid

        return jsonify({"success": True, "id": assoc_id}), 201
    except Exception as e:
        return jsonify({"error": f"创建关联失败: {str(e)}"}), 500


@api.route("/api/fc-cp-association", methods=["DELETE"])
@login_required
def delete_fc_cp_association():
    """删除 FC-CP 关联"""
    assoc_id = request.args.get("id", type=int)

    if not assoc_id:
        return jsonify({"error": "缺少关联 ID"}), 400

    # 获取项目 ID 来确定数据库
    project_id = request.args.get("project_id", type=int)
    if not project_id:
        return jsonify({"error": "缺少 project_id"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    cursor.execute("DELETE FROM fc_cp_association WHERE id=?", (assoc_id,))
    conn.commit()

    return jsonify({"success": True})


@api.route("/api/fc-cp-association/import", methods=["POST"])
@login_required
def import_fc_cp_association():
    """导入 FC-CP 关联 (CSV)"""
    project_id = request.args.get("project_id", type=int)

    if not project_id:
        return jsonify({"error": "缺少 project_id"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    # 获取 CSV 数据
    csv_data = request.json.get("csv_data", [])
    if not csv_data or len(csv_data) < 2:
        return jsonify({"error": "CSV 数据为空或格式错误"}), 400

    headers = csv_data[0]
    header_map = {}
    for idx, header in enumerate(headers):
        if header:
            header_map[header.strip()] = idx

    # 检查必填字段
    required_fields = ["cp_feature", "cp_sub_feature", "cp_cover_point", "fc_covergroup", "fc_coverpoint", "fc_bin_name"]
    for field in required_fields:
        if field not in header_map:
            return jsonify({"error": f"缺少必填字段: {field}"}), 400

    conn = get_db(project["name"])
    cursor = conn.cursor()

    imported = 0
    errors = []
    created_by = session.get(SESSION_USERNAME_KEY, "unknown")

    for row_idx, row in enumerate(csv_data[1:], start=2):
        try:
            cp_feature = (
                row[header_map.get("cp_feature", 0)]
                if header_map.get("cp_feature", 0) < len(row)
                else None
            )
            cp_sub_feature = (
                row[header_map.get("cp_sub_feature", 1)]
                if header_map.get("cp_sub_feature", 1) < len(row)
                else None
            )
            cp_cover_point = (
                row[header_map.get("cp_cover_point", 2)]
                if header_map.get("cp_cover_point", 2) < len(row)
                else None
            )
            fc_covergroup = (
                row[header_map.get("fc_covergroup", 3)]
                if header_map.get("fc_covergroup", 3) < len(row)
                else None
            )
            fc_coverpoint = (
                row[header_map.get("fc_coverpoint", 4)]
                if header_map.get("fc_coverpoint", 4) < len(row)
                else None
            )
            fc_bin_name = (
                row[header_map.get("fc_bin_name", 5)]
                if header_map.get("fc_bin_name", 5) < len(row)
                else None
            )

            if not all([cp_feature, cp_cover_point, fc_covergroup, fc_coverpoint, fc_bin_name]):
                errors.append(f"第{row_idx}行: 必填字段缺失")
                continue

            # 查找 CP ID (使用 feature + sub_feature + cover_point)
            cursor.execute(
                "SELECT id FROM cover_point WHERE feature=? AND sub_feature=? AND cover_point=?",
                (cp_feature, cp_sub_feature or "", cp_cover_point)
            )
            cp_row = cursor.fetchone()
            if not cp_row:
                errors.append(f'第{row_idx}行: CP "{cp_feature}/{cp_sub_feature}/{cp_cover_point}" 不存在')
                continue
            cp_id = cp_row[0]

            # 查找 FC ID (使用 covergroup + coverpoint + bin_name)
            cursor.execute(
                "SELECT id FROM functional_coverage WHERE covergroup=? AND coverpoint=? AND bin_name=?",
                (fc_covergroup, fc_coverpoint, fc_bin_name)
            )
            fc_row = cursor.fetchone()
            if not fc_row:
                errors.append(f'第{row_idx}行: FC "{fc_covergroup}/{fc_coverpoint}/{fc_bin_name}" 不存在')
                continue
            fc_id = fc_row[0]

            # 检查关联是否已存在
            cursor.execute("SELECT id FROM fc_cp_association WHERE cp_id=? AND fc_id=?", (cp_id, fc_id))
            if cursor.fetchone():
                errors.append(f'第{row_idx}行: 关联已存在')
                continue

            cursor.execute(
                """
                INSERT INTO fc_cp_association (project_id, cp_id, fc_id, created_by, created_at)
                VALUES (?, ?, ?, ?, datetime('now'))
                """,
                (project_id, cp_id, fc_id, created_by)
            )

            imported += 1
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")

    if imported > 0:
        conn.commit()

    return jsonify({"success": True, "imported": imported, "failed": len(errors), "errors": errors})


# ============ 统计 ============


@api.route("/api/stats", methods=["GET"])
def get_stats():
    """获取统计数据"""
    project_id = request.args.get("project_id", type=int)

    if not project_id:
        return jsonify(
            {
                "total_cp": 0,
                "total_tc": 0,
                "open_tc": 0,
                "coded_tc": 0,
                "fail_tc": 0,
                "pass_tc": 0,
                "coverage": "0%",
            }
        )

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify(
            {
                "total_cp": 0,
                "total_tc": 0,
                "open_tc": 0,
                "coded_tc": 0,
                "fail_tc": 0,
                "pass_tc": 0,
                "coverage": "0%",
            }
        )

    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 检查数据库表是否存在
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = {row[0] for row in cursor.fetchall()}

    if "cover_point" not in tables or "test_case" not in tables:
        return jsonify(
            {
                "total_cp": 0,
                "total_tc": 0,
                "open_tc": 0,
                "coded_tc": 0,
                "fail_tc": 0,
                "pass_tc": 0,
                "coverage": "0%",
            }
        )

    # CP 统计
    cursor.execute("SELECT COUNT(*) FROM cover_point")
    total_cp = cursor.fetchone()[0]

    # TC 统计（REMOVED 不计入 Total）
    cursor.execute('SELECT COUNT(*) FROM test_case WHERE status != "REMOVED"')
    total_tc = cursor.fetchone()[0]

    cursor.execute('SELECT COUNT(*) FROM test_case WHERE status="OPEN" AND status != "REMOVED"')
    open_tc = cursor.fetchone()[0]

    cursor.execute('SELECT COUNT(*) FROM test_case WHERE status="CODED"')
    coded_tc = cursor.fetchone()[0]

    cursor.execute('SELECT COUNT(*) FROM test_case WHERE status="FAIL"')
    fail_tc = cursor.fetchone()[0]

    cursor.execute('SELECT COUNT(*) FROM test_case WHERE status="PASS"')
    pass_tc = cursor.fetchone()[0]

    # Pass Rate 计算: PASS / Total * 100%
    pass_rate = 0
    if total_tc > 0:
        pass_rate = round(pass_tc / total_tc * 100, 1)

    # 计算覆盖率（只统计非 REMOVED 的 TC）
    coverage = 0
    if total_cp > 0:
        cursor.execute("SELECT id FROM cover_point")
        cp_ids = [row[0] for row in cursor.fetchall()]

        total_progress = 0
        for cp_id in cp_ids:
            cursor.execute("SELECT tc_id FROM tc_cp_connections WHERE cp_id=?", (cp_id,))
            tc_ids = [r[0] for r in cursor.fetchall()]

            if tc_ids:
                placeholders = ",".join(["?"] * len(tc_ids))
                cursor.execute(
                    f'SELECT COUNT(*) FROM test_case WHERE id IN ({placeholders}) AND status="PASS"',
                    tc_ids,
                )
                passed = cursor.fetchone()[0]
                total_progress += (passed / len(tc_ids)) * 100

        coverage = round(total_progress / total_cp, 1)

    return jsonify(
        {
            "total_cp": total_cp,
            "total_tc": total_tc,
            "open_tc": open_tc,
            "coded_tc": coded_tc,
            "fail_tc": fail_tc,
            "pass_tc": pass_tc,
            "pass_rate": f"{pass_rate}%",
            "coverage": f"{coverage}%",
        }
    )


# ============ 导入导出功能 ============

import base64
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment


@api.route("/api/import/template", methods=["GET"])
@guest_required
def get_import_template():
    """下载导入模板"""
    template_type = request.args.get("type", "cp")

    wb = Workbook()
    ws = wb.active

    if template_type == "cp":
        # CP 模板
        headers = ["Feature", "Sub-Feature", "Cover Point", "Cover Point Details", "Comments"]
        ws.append(headers)
    elif template_type == "tc":
        # TC 模板
        headers = [
            "TestBench",
            "Category",
            "Owner",
            "Test Name",
            "Scenario Details",
            "Checker Details",
            "Coverage Details",
            "Comments",
        ]
        ws.append(headers)
    elif template_type == "connection":
        # TC-CP 关联导入模板
        headers = ["Test Case", "Cover Point"]
        ws.append(headers)
    else:
        return jsonify({"error": "无效的模板类型"}), 400

    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 设置列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{template_type.upper()}_import_template.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
@guest_required

@guest_required

@api.route("/api/import", methods=["POST"])
@guest_required
def import_data():
    """执行导入"""
    data = request.json
    project_id = data.get("project_id")
    import_type = data.get("type")  # 'cp' or 'tc'
    file_data = data.get("file_data")  # base64 encoded

    if not project_id or not import_type or not file_data:
        return jsonify({"error": "缺少必要参数"}), 400

    if import_type not in ["cp", "tc", "connection"]:
        return jsonify({"error": "无效的导入类型"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    try:
        # 解码 base64 文件内容
        file_content = base64.b64decode(file_data)

        # 根据文件内容判断类型（检查 BOM 或内容）
        # Excel 文件以 PK 开头（ZIP 格式），CSV 文件是纯文本
        is_csv = not file_content[:2] == b"PK"

        if is_csv:
            # 使用 csv 模块读取（需要文本模式）
            import csv

            # file_content 已经是 base64 解码后的 bytes，转为 string
            csv_reader = csv.reader(io.StringIO(file_content.decode("utf-8")))
            rows = list(csv_reader)
            if not rows:
                return jsonify({"error": "CSV 文件为空"}), 400
            headers = rows[0] if rows else []
            # 传递整个 rows（包含 header），在 import_cp/import_tc 中处理
            csv_data = rows
            csv_headers = headers
        else:
            # 使用 openpyxl 读取 Excel (使用临时文件方式解决 BytesIO 问题)
            import tempfile
            import os

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(file_content)
                tmp_path = tmp.name

            try:
                wb = load_workbook(tmp_path)
                ws = wb.active
                csv_data = None
                csv_headers = None
            finally:
                os.unlink(tmp_path)  # 清理临时文件

        # 获取表头
        if csv_headers is not None:
            headers = csv_headers
        else:
            headers = [cell.value for cell in ws[1]]

        if import_type == "cp":
            return import_cp(project, ws if not is_csv else None, headers, is_csv, csv_data)
        elif import_type == "tc":
            return import_tc(project, ws if not is_csv else None, headers, is_csv, csv_data)
        else:
            return import_connections(project, ws if not is_csv else None, headers, is_csv, csv_data)

    except Exception as e:
        return jsonify({"error": f"导入失败: {str(e)}"}), 400


def import_cp(project, ws, headers, is_csv=False, csv_data=None):
    """导入 CP 数据"""
    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 解析表头
    header_map = {}
    for idx, header in enumerate(headers):
        if header:
            header_map[header.strip()] = idx  # 0-based 索引

    # 检查必填字段
    required_fields = ["Feature", "Cover Point"]
    for field in required_fields:
        if field not in header_map:
            return jsonify({"error": f"缺少必填字段: {field}"}), 400

    imported = 0
    errors = []

    # 根据类型选择读取方式
    if is_csv and csv_data:
        # CSV: csv_data 是所有行的列表，第一行是 header
        for row_idx, row in enumerate(csv_data[1:], start=2):  # 跳过 header
            try:
                feature = (
                    row[header_map.get("Feature", 0)]
                    if header_map.get("Feature", 0) < len(row)
                    else None
                )
                cover_point = (
                    row[header_map.get("Cover Point", 2)]
                    if header_map.get("Cover Point", 0) < len(row)
                    else None
                )

                if not feature or not cover_point:
                    errors.append(f"第{row_idx}行: 必填字段缺失")
                    continue

                # 获取 sub_feature 用于重名检查 (v0.11.0)
                sub_feature = (
                    row[header_map.get("Sub-Feature", 1)]
                    if header_map.get("Sub-Feature", 1) < len(row)
                    else ""
                )

                # v0.11.0: 检查重名使用 feature + sub_feature + cover_point 组合
                cursor.execute(
                    "SELECT id FROM cover_point WHERE feature=? AND sub_feature=? AND cover_point=?",
                    (feature, sub_feature, cover_point)
                )
                if cursor.fetchone():
                    errors.append(f'第{row_idx}行: CP "{feature}/{sub_feature}/{cover_point}" 已存在')
                    continue
                cover_point_details = (
                    row[header_map.get("Cover Point Details", 3)]
                    if header_map.get("Cover Point Details", 3) < len(row)
                    else ""
                )
                # Priority 字段 (v0.9.0 修复)
                priority = (
                    row[header_map.get("Priority", 5)]
                    if header_map.get("Priority", 5) < len(row)
                    else "P0"
                )
                # 验证 Priority 值合法性
                if priority not in ["P0", "P1", "P2"]:
                    priority = "P0"
                comments = (
                    row[header_map.get("Comments", 6)]
                    if header_map.get("Comments", 6) < len(row)
                    else ""
                )

                cursor.execute(
                    """
                    INSERT INTO cover_point (project_id, feature, sub_feature, cover_point, cover_point_details, comments, priority, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                    (
                        project["id"],
                        feature,
                        sub_feature,
                        cover_point,
                        cover_point_details,
                        comments,
                        priority,
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )

                imported += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
    else:
        # Excel: 使用 ws.cell() (1-based 索引)
        for row_idx in range(2, ws.max_row + 1):
            try:
                feature = ws.cell(row_idx, header_map.get("Feature", 0) + 1).value
                cover_point = ws.cell(row_idx, header_map.get("Cover Point", 0) + 1).value

                if not feature or not cover_point:
                    errors.append(f"第{row_idx}行: 必填字段缺失")
                    continue

                # 获取 sub_feature 用于重名检查 (v0.11.0)
                sub_feature = ws.cell(row_idx, header_map.get("Sub-Feature", 0) + 1).value or ""

                # v0.11.0: 检查重名使用 feature + sub_feature + cover_point 组合
                cursor.execute(
                    "SELECT id FROM cover_point WHERE feature=? AND sub_feature=? AND cover_point=?",
                    (feature, sub_feature, cover_point)
                )
                if cursor.fetchone():
                    errors.append(f'第{row_idx}行: CP "{feature}/{sub_feature}/{cover_point}" 已存在')
                    continue

                # 获取其他字段
                cover_point_details = (
                    ws.cell(row_idx, header_map.get("Cover Point Details", 0) + 1).value or ""
                )
                # Priority 字段 (v0.9.0 修复)
                priority_cell = ws.cell(row_idx, header_map.get("Priority", 0) + 1).value
                priority = priority_cell if priority_cell in ["P0", "P1", "P2"] else "P0"
                comments = ws.cell(row_idx, header_map.get("Comments", 0) + 1).value or ""

                cursor.execute(
                    """
                    INSERT INTO cover_point (project_id, feature, sub_feature, cover_point, cover_point_details, comments, priority, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                    (
                        project["id"],
                        feature,
                        sub_feature,
                        cover_point,
                        cover_point_details,
                        comments,
                        priority,
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )

                imported += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")

    if imported > 0:
        conn.commit()

    return jsonify({"success": True, "imported": imported, "failed": len(errors), "errors": errors})


def import_tc(project, ws, headers, is_csv=False, csv_data=None):
    """导入 TC 数据"""
    conn = get_db(project["name"])
    cursor = conn.cursor()

    # 解析表头
    header_map = {}
    for idx, header in enumerate(headers):
        if header:
            header_map[header.strip()] = idx  # 0-based 索引

    # 检查必填字段
    required_fields = ["TestBench", "Test Name"]
    for field in required_fields:
        if field not in header_map:
            return jsonify({"error": f"缺少必填字段: {field}"}), 400

    imported = 0
    errors = []

    # 根据类型选择读取方式
    if is_csv and csv_data:
        # CSV: csv_data 是所有行的列表，第一行是 header
        for row_idx, row in enumerate(csv_data[1:], start=2):  # 跳过 header
            try:
                testbench = (
                    row[header_map.get("TestBench", 0)]
                    if header_map.get("TestBench", 0) < len(row)
                    else None
                )
                test_name = (
                    row[header_map.get("Test Name", 3)]
                    if header_map.get("Test Name", 3) < len(row)
                    else None
                )

                if not testbench or not test_name:
                    errors.append(f"第{row_idx}行: 必填字段缺失")
                    continue

                # v0.11.0: 检查重名使用 testbench + test_name 组合
                cursor.execute(
                    "SELECT id FROM test_case WHERE testbench=? AND test_name=?",
                    (testbench, test_name)
                )
                if cursor.fetchone():
                    errors.append(f'第{row_idx}行: TC "{testbench}/{test_name}" 已存在')
                    continue

                # 获取其他字段
                category = (
                    row[header_map.get("Category", 1)]
                    if header_map.get("Category", 1) < len(row)
                    else ""
                )
                owner = (
                    row[header_map.get("Owner", 2)] if header_map.get("Owner", 2) < len(row) else ""
                )
                scenario_details = (
                    row[header_map.get("Scenario Details", 4)]
                    if header_map.get("Scenario Details", 4) < len(row)
                    else ""
                )
                checker_details = (
                    row[header_map.get("Checker Details", 5)]
                    if header_map.get("Checker Details", 5) < len(row)
                    else ""
                )
                coverage_details = (
                    row[header_map.get("Coverage Details", 6)]
                    if header_map.get("Coverage Details", 6) < len(row)
                    else ""
                )
                comments = (
                    row[header_map.get("Comments", 7)]
                    if header_map.get("Comments", 7) < len(row)
                    else ""
                )

                cursor.execute(
                    """
                    INSERT INTO test_case (
                        project_id, dv_milestone, testbench, category, owner, test_name,
                        scenario_details, checker_details, coverage_details, comments,
                        priority, status, created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'OPEN', ?)
                """,
                    (
                        project["id"],
                        "DV1.0",
                        testbench,
                        category,
                        owner,
                        test_name,
                        scenario_details,
                        checker_details,
                        coverage_details,
                        comments,
                        "P0",
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )

                imported += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
    else:
        # Excel: 使用 ws.cell() (1-based 索引)
        for row_idx in range(2, ws.max_row + 1):
            try:
                testbench = ws.cell(row_idx, header_map.get("TestBench", 0) + 1).value
                test_name = ws.cell(row_idx, header_map.get("Test Name", 0) + 1).value

                if not testbench or not test_name:
                    errors.append(f"第{row_idx}行: 必填字段缺失")
                    continue

                # v0.11.0: 检查重名使用 testbench + test_name 组合
                cursor.execute(
                    "SELECT id FROM test_case WHERE testbench=? AND test_name=?",
                    (testbench, test_name)
                )
                if cursor.fetchone():
                    errors.append(f'第{row_idx}行: TC "{testbench}/{test_name}" 已存在')
                    continue

                # 获取其他字段
                category = ws.cell(row_idx, header_map.get("Category", 0) + 1).value or ""
                owner = ws.cell(row_idx, header_map.get("Owner", 0) + 1).value or ""
                scenario_details = (
                    ws.cell(row_idx, header_map.get("Scenario Details", 0) + 1).value or ""
                )
                checker_details = (
                    ws.cell(row_idx, header_map.get("Checker Details", 0) + 1).value or ""
                )
                coverage_details = (
                    ws.cell(row_idx, header_map.get("Coverage Details", 0) + 1).value or ""
                )
                comments = ws.cell(row_idx, header_map.get("Comments", 0) + 1).value or ""

                cursor.execute(
                    """
                    INSERT INTO test_case (
                        project_id, dv_milestone, testbench, category, owner, test_name,
                        scenario_details, checker_details, coverage_details, comments,
                        priority, status, created_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'OPEN', ?)
                """,
                    (
                        project["id"],
                        "DV1.0",
                        testbench,
                        category,
                        owner,
                        test_name,
                        scenario_details,
                        checker_details,
                        coverage_details,
                        comments,
                        "P0",
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ),
                )

                imported += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")

    if imported > 0:
        conn.commit()

    return jsonify({"success": True, "imported": imported, "failed": len(errors), "errors": errors})


def import_connections(project, ws, headers, is_csv=False, csv_data=None):
    """导入 TC-CP 关联 (v0.9.1 新增)
    
    CSV 格式:
    Test Case,Cover Point
    TC名称1,CP名称1
    TC名称2,CP名称2
    
    支持多对多关联：一个 TC 可以关联多个 CP，一个 CP 可以被多个 TC 关联
    """
    conn = get_db(project["name"])
    cursor = conn.cursor()
    
    # 解析表头
    header_map = {}
    for idx, header in enumerate(headers):
        if header:
            header_map[header.strip()] = idx
    
    # 必填字段
    required_fields = ["Test Case", "Cover Point"]
    for field in required_fields:
        if field not in header_map:
            return jsonify({"error": f"缺少必填字段: {field}"}), 400
    
    imported = 0
    skipped = 0
    errors = []
    
    # 获取数据行
    if is_csv and csv_data:
        data_rows = csv_data[1:]  # 跳过 header
    else:
        # Excel 格式
        data_rows = []
        for row_idx in range(2, ws.max_row + 1):
            row = [ws.cell(row_idx, col_idx + 1).value for col_idx in range(len(headers))]
            data_rows.append(row)
    
    for row_idx, row in enumerate(data_rows, start=2):
        try:
            # 安全获取 TC 名称 (兼容空值和 None)
            tc_idx = header_map.get("Test Case", 0)
            tc_name = ""
            if tc_idx < len(row):
                tc_val = row[tc_idx]
                if tc_val is not None:
                    tc_name = str(tc_val).strip()
            
            # 安全获取 CP 名称 (兼容空值和 None)
            cp_idx = header_map.get("Cover Point", 1)
            cp_name = ""
            if cp_idx < len(row):
                cp_val = row[cp_idx]
                if cp_val is not None:
                    cp_name = str(cp_val).strip()
            
            if not tc_name or not cp_name:
                errors.append(f"第{row_idx}行: TC 或 CP 名称为空")
                continue
            
            # 查询 TC ID
            cursor.execute(
                "SELECT id FROM test_case WHERE test_name = ?", (tc_name,)
            )
            tc_result = cursor.fetchone()
            if not tc_result:
                errors.append(f"第{row_idx}行: TC '{tc_name}' 不存在")
                continue
            tc_id = tc_result["id"]
            
            # 查询 CP ID
            cursor.execute(
                "SELECT id FROM cover_point WHERE cover_point = ?", (cp_name,)
            )
            cp_result = cursor.fetchone()
            if not cp_result:
                errors.append(f"第{row_idx}行: CP '{cp_name}' 不存在")
                continue
            cp_id = cp_result["id"]
            
            # 检查关联是否已存在
            cursor.execute(
                "SELECT 1 FROM tc_cp_connections WHERE tc_id = ? AND cp_id = ?",
                (tc_id, cp_id)
            )
            if cursor.fetchone():
                skipped += 1
                continue
            
            # 插入关联
            cursor.execute(
                "INSERT INTO tc_cp_connections (tc_id, cp_id) VALUES (?, ?)",
                (tc_id, cp_id)
            )
            imported += 1
            
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")
    
    if imported > 0:
        conn.commit()
    
    return jsonify({
        "success": True,
        "imported": imported,
        "skipped": skipped,
        "failed": len(errors),
        "errors": errors
    })


@api.route("/api/export", methods=["GET"])
@guest_required
def export_data():
    """导出数据 - 必须在静态文件路由之前定义"""
    project_id = request.args.get("project_id", type=int)
    export_type = request.args.get("type")  # 'cp' or 'tc'
    export_format = request.args.get("format", "xlsx")  # 'xlsx' or 'csv'

    if not project_id or not export_type:
        return jsonify({"error": "缺少必要参数"}), 400

    if export_type not in ["cp", "tc"]:
        return jsonify({"error": "无效的导出类型"}), 400

    projects = load_projects()
    project = next((p for p in projects if p["id"] == project_id), None)

    if not project:
        return jsonify({"error": "项目不存在"}), 404

    conn = get_db(project["name"])
    cursor = conn.cursor()

    if export_type == "cp":
        cursor.execute("SELECT * FROM cover_point ORDER BY id")
        rows = cursor.fetchall()

        # CP 导出字段
        export_headers = [
            "Feature",
            "Sub-Feature",
            "Cover Point",
            "Cover Point Details",
            "Priority",
            "Comments",
            "Created At",
        ]
        export_fields = [
            "feature",
            "sub_feature",
            "cover_point",
            "cover_point_details",
            "priority",
            "comments",
            "created_at",
        ]
    else:
        cursor.execute("SELECT * FROM test_case ORDER BY id")
        rows = cursor.fetchall()

        # TC 导出字段
        export_headers = [
            "ID",
            "DV Milestone",
            "TestBench",
            "Category",
            "Owner",
            "Test Name",
            "Scenario Details",
            "Checker Details",
            "Coverage Details",
            "Status",
            "Comments",
            "Created At",
            "Coded Date",
            "Fail Date",
            "Pass Date",
            "Removed Date",
            "Target Date",
        ]
        export_fields = [
            "id",
            "dv_milestone",
            "testbench",
            "category",
            "owner",
            "test_name",
            "scenario_details",
            "checker_details",
            "coverage_details",
            "status",
            "comments",
            "created_at",
            "coded_date",
            "fail_date",
            "pass_date",
            "removed_date",
            "target_date",
        ]

    if export_format == "xlsx":
        # 导出为 Excel
        wb = Workbook()
        ws = wb.active

        # 写入表头
        ws.append(export_headers)

        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 写入数据
        for row in rows:
            row_data = []
            for field in export_fields:
                row_data.append(row[field] or "")
            ws.append(row_data)

        # 设置列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"{project['name']}_{export_type.upper()}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        # 导出为 CSV
        import csv

        output = io.StringIO()
        writer = csv.writer(output)

        # 写入表头
        writer.writerow(export_headers)

        # 写入数据
        for row in rows:
            row_data = []
            for field in export_fields:
                row_data.append(str(row[field] or ""))
            writer.writerow(row_data)

        output.seek(0)

        filename = (
            f"{project['name']}_{export_type.upper()}_{datetime.now().strftime('%Y%m%d')}.csv"
        )

        # 返回 CSV - 对文件名进行 URL 编码以支持中文
        from flask import Response

        return Response(
            output.getvalue(),
            mimetype="text/csv;charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={quote(filename)}"},
        )


# ============ 认证 API ============

@api.route("/api/auth/login", methods=["POST"])
def login():
    """用户登录"""
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Bad Request", "message": "无效的请求数据"}), 400
    
    username = data.get("username", "").strip()
    password = data.get("password", "")
    
    if not username or not password:
        return jsonify({"error": "Bad Request", "message": "用户名和密码不能为空"}), 400
    
    # 检查暴力破解
    allowed, message = auth.check_login_attempts(username)
    if not allowed:
        return jsonify({"error": "Too Many Requests", "message": message}), 429
    
    # 获取用户
    user = auth.get_user_by_username(username)
    
    if not user:
        auth.record_failed_login(username)
        return jsonify({"error": "Unauthorized", "message": "用户名或密码错误"}), 401
    
    # 检查用户是否启用
    if not user.get("is_active"):
        return jsonify({"error": "Forbidden", "message": "账户已被禁用"}), 403
    
    # 验证密码
    if not auth.verify_password(password, user.get("password_hash")):
        auth.record_failed_login(username)
        return jsonify({"error": "Unauthorized", "message": "用户名或密码错误"}), 401
    
    # 登录成功，清除失败记录
    auth.clear_login_attempts(username)
    
    # 更新最后登录时间
    auth.update_user(user["id"], last_login=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    # 设置 Session
    session[SESSION_USER_KEY] = user["id"]
    session[SESSION_USERNAME_KEY] = user["username"]
    session[SESSION_ROLE_KEY] = user["role"]
    session.permanent = True
    
    return jsonify({
        "success": True,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "role": user["role"],
            "must_change_password": user.get("must_change_password", 0) == 1
        }
    })


@api.route("/api/auth/guest-login", methods=["POST"])
def guest_login():
    """访客登录"""
    # 检查暴力破解
    allowed, message = auth.check_login_attempts("guest")
    if not allowed:
        return jsonify({"error": "Too Many Requests", "message": message}), 429
    
    # 获取 guest 用户
    user = auth.get_user_by_username("guest")
    
    if not user:
        return jsonify({"error": "Not Found", "message": "访客账户未配置"}), 404
    
    # 检查用户是否启用
    if not user.get("is_active"):
        return jsonify({"error": "Forbidden", "message": "访客登录已禁用"}), 403
    
    # 登录成功
    auth.clear_login_attempts("guest")
    
    # 更新最后登录时间
    auth.update_user(user["id"], last_login=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    # 设置 Session
    session[SESSION_USER_KEY] = user["id"]
    session[SESSION_USERNAME_KEY] = user["username"]
    session[SESSION_ROLE_KEY] = user["role"]
    session.permanent = True
    
    return jsonify({
        "success": True,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "role": user["role"]
        }
    })


@api.route("/api/auth/logout", methods=["POST"])
@login_required
def logout():
    """退出登录"""
    session.clear()
    return jsonify({"success": True, "message": "已退出登录"})


@api.route("/api/auth/me", methods=["GET"])
@login_required
def get_current_user():
    """获取当前用户信息"""
    user_id = session.get(SESSION_USER_KEY)
    user = auth.get_user_by_id(user_id)
    
    if not user:
        session.clear()
        return jsonify({"error": "Unauthorized", "message": "用户不存在"}), 401
    
    return jsonify({
        "id": user["id"],
        "username": user["username"],
        "role": user["role"],
        "must_change_password": user.get("must_change_password", 0) == 1
    })


@api.route("/api/auth/password", methods=["PATCH"])
@login_required
def change_password():
    """修改当前用户密码"""
    data = request.get_json()
    new_password = data.get("password", "") if data else ""

    # 验证密码长度
    if not new_password:
        return jsonify({"error": "Bad Request", "message": "密码不能为空"}), 400

    if len(new_password) < 6:
        return jsonify({"error": "Bad Request", "message": "密码长度至少6位"}), 400

    # 获取当前用户ID
    user_id = session.get(SESSION_USER_KEY)
    user = auth.get_user_by_id(user_id)

    if not user:
        return jsonify({"error": "Not Found", "message": "用户不存在"}), 404

    # guest 账户不能修改密码
    if user["role"] == "guest":
        return jsonify({"error": "Forbidden", "message": "访客账户无密码"}), 403

    # 更新密码
    password_hash = auth.hash_password(new_password)
    auth.update_user(user_id, password_hash=password_hash, must_change_password=0)

    return jsonify({"success": True, "message": "密码修改成功"})


# ============ 用户管理 API ============

@api.route("/api/users", methods=["GET"])
@admin_required
def get_users():
    """获取用户列表（仅管理员）"""
    users = auth.get_all_users()
    return jsonify(users)


@api.route("/api/users", methods=["POST"])
@admin_required
def create_user():
    """创建用户（仅管理员）"""
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Bad Request", "message": "无效的请求数据"}), 400
    
    username = data.get("username", "").strip()
    password = data.get("password", "")
    role = data.get("role", "user")
    
    if not username:
        return jsonify({"error": "Bad Request", "message": "用户名不能为空"}), 400
    
    if role not in ['user', 'guest']:
        return jsonify({"error": "Bad Request", "message": "无效的角色"}), 400
    
    # guest 账户不能有密码
    if role == 'guest' and password:
        return jsonify({"error": "Bad Request", "message": "访客账户不能设置密码"}), 400
    
    # 检查用户名是否已存在
    existing = auth.get_user_by_username(username)
    if existing:
        return jsonify({"error": "Conflict", "message": "用户名已存在"}), 409
    
    try:
        user_id = auth.create_user(username, password, role)
        return jsonify({"success": True, "id": user_id, "username": username, "role": role}), 201
    except Exception as e:
        return jsonify({"error": "Internal Error", "message": str(e)}), 500


@api.route("/api/users/batch", methods=["POST"])
@admin_required
def batch_create_users():
    """批量创建用户（仅管理员）- 仅支持创建普通用户"""
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Bad Request", "message": "无效的请求数据"}), 400
    
    users = data.get("users", [])
    
    if not users or not isinstance(users, list):
        return jsonify({"error": "Bad Request", "message": "请提供用户列表"}), 400
    
    # 限制单次批量创建数量
    if len(users) > 50:
        return jsonify({"error": "Bad Request", "message": "单次最多创建 50 个用户"}), 400
    
    default_password = "123456"
    results = []
    created_count = 0
    failed_count = 0
    
    for user_data in users:
        username = user_data.get("username", "").strip() if isinstance(user_data, dict) else str(user_data).strip()
        
        if not username:
            results.append({"username": username or "(空)", "status": "failed", "error": "用户名不能为空"})
            failed_count += 1
            continue
        
        # 安全限制：仅支持创建普通用户 role=user
        role = "user"
        
        # 检查用户名是否已存在
        existing = auth.get_user_by_username(username)
        if existing:
            results.append({"username": username, "status": "failed", "error": "用户名已存在"})
            failed_count += 1
            continue
        
        try:
            user_id = auth.create_user(username, default_password, role)
            results.append({"username": username, "status": "created", "id": user_id})
            created_count += 1
        except Exception as e:
            results.append({"username": username, "status": "failed", "error": str(e)})
            failed_count += 1
    
    return jsonify({
        "success": failed_count == 0,
        "created": created_count,
        "failed": failed_count,
        "password": default_password,
        "results": results
    }), 200


@api.route("/api/users/<int:user_id>", methods=["PATCH"])
@admin_required
def update_user(user_id):
    """更新用户（仅管理员）"""
    data = request.get_json()
    
    if not data:
        return jsonify({"error": "Bad Request", "message": "无效的请求数据"}), 400
    
    # 获取现有用户
    user = auth.get_user_by_id(user_id)
    if not user:
        return jsonify({"error": "Not Found", "message": "用户不存在"}), 404
    
    # 不能修改 admin 和 guest 的角色
    if user["username"] in ['admin', 'guest']:
        if "role" in data:
            return jsonify({"error": "Forbidden", "message": "不能修改系统用户的角色"}), 403
    
    # 构建更新字段
    updates = {}
    if "is_active" in data:
        updates["is_active"] = 1 if data["is_active"] else 0
    
    if updates:
        auth.update_user(user_id, **updates)
    
    return jsonify({"success": True})


@api.route("/api/users/<int:user_id>", methods=["DELETE"])
@admin_required
def delete_user(user_id):
    """删除用户（仅管理员）"""
    # 获取现有用户
    user = auth.get_user_by_id(user_id)
    if not user:
        return jsonify({"error": "Not Found", "message": "用户不存在"}), 404
    
    # 不能删除 admin 和 guest
    if user["username"] in ['admin', 'guest']:
        return jsonify({"error": "Forbidden", "message": "无法删除系统用户"}), 403
    
    # 不能删除自己
    current_user_id = session.get(SESSION_USER_KEY)
    if current_user_id == user_id:
        return jsonify({"error": "Forbidden", "message": "不能删除自己"}), 403
    
    success, message = auth.delete_user(user_id)
    
    if not success:
        return jsonify({"error": "Forbidden", "message": message}), 403
    
    return jsonify({"success": True, "message": message})


@api.route("/api/users/<int:user_id>/reset-password", methods=["POST"])
@admin_required
def reset_password(user_id):
    """重置用户密码（仅管理员）"""
    data = request.get_json()
    new_password = data.get("new_password", "") if data else ""
    
    # 获取现有用户
    user = auth.get_user_by_id(user_id)
    if not user:
        return jsonify({"error": "Not Found", "message": "用户不存在"}), 404
    
    # guest 账户不能重置密码
    if user["role"] == "guest":
        return jsonify({"error": "Forbidden", "message": "访客账户无密码"}), 403
    
    if not new_password:
        return jsonify({"error": "Bad Request", "message": "密码不能为空"}), 400
    
    # 重置密码
    password_hash = auth.hash_password(new_password)
    auth.update_user(user_id, password_hash=password_hash, must_change_password=0)
    
    return jsonify({"success": True, "message": "密码已重置"})


# ============ 用户反馈 API ============

@api.route("/api/feedback", methods=["POST"])
@login_required
def submit_feedback():
    """提交用户反馈"""
    data = request.get_json()

    if not data:
        return jsonify({"error": "Bad Request", "message": "无效的请求数据"}), 400

    feedback_type = data.get("type", "").strip()
    title = data.get("title", "").strip()
    description = data.get("description", "").strip()

    # 验证必填字段
    if not feedback_type:
        return jsonify({"error": "Bad Request", "message": "请选择反馈类型"}), 400
    if feedback_type not in ["bug", "feature", "optimization"]:
        return jsonify({"error": "Bad Request", "message": "无效的反馈类型"}), 400
    if not title:
        return jsonify({"error": "Bad Request", "message": "请输入反馈标题"}), 400
    if not description:
        return jsonify({"error": "Bad Request", "message": "请输入反馈描述"}), 400

    # 获取当前用户
    username = session.get(SESSION_USERNAME_KEY, "anonymous")

    # 生成反馈 ID
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    feedback_id = f"FEEDBACK_{timestamp}"

    # 创建反馈数据
    feedback_data = {
        "id": feedback_id,
        "type": feedback_type,
        "title": title,
        "description": description,
        "created_at": datetime.now().isoformat() + "Z",
        "user": username
    }

    # 保存到文件
    feedback_dir = os.path.join(current_app.config["DATA_DIR"], "feedbacks")
    os.makedirs(feedback_dir, exist_ok=True)

    feedback_file = os.path.join(feedback_dir, f"{feedback_id}.json")
    try:
        with open(feedback_file, "w", encoding="utf-8") as f:
            json.dump(feedback_data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return jsonify({"error": "Server Error", "message": f"保存反馈失败: {str(e)}"}), 500

    return jsonify({
        "success": True,
        "message": "反馈提交成功",
        "data": feedback_data
    })


# ============ 静态文件路由 - 必须放在所有 API 路由之后 ============

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


@api.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


# 注意：这个路由必须放在所有 API 路由之后
# 它使用 path 变量，会匹配其他未匹配的路由
